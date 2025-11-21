# -*- coding: utf-8 -*-
"""
Skrypt tworzący rozszerzony kalkulator opłacalności zakupu nieruchomości w Szwajcarii.
Wersja z harmonogramami rocznymi i miesięcznymi oraz dobrowolną amortyzacją H1.

INSTRUKCJA UŻYCIA (Linux Mint):
1. Zainstaluj openpyxl: 
   sudo apt install python3-openpyxl
   
2. Uruchom skrypt:
   python3 build_kalkulator_nieruchomosc_ch.py
   
3. Plik kalkulator_nieruchomosc_CH.xlsx zostanie utworzony w bieżącym katalogu

4. Otwórz plik w LibreOffice Calc:
   libreoffice --calc kalkulator_nieruchomosc_CH.xlsx

AUTOR: GitHub Copilot
DATA: 2025-11-20
UŻYTKOWNIK: matin-chuj
SYSTEM: Linux Mint
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00


def set_cell_style(cell, font_bold=False, font_size=11, bg_color=None, 
                   border=True, number_format=None, alignment='left'):
    """Pomocnicza funkcja do ustawiania stylu komórki."""
    if font_bold or font_size != 11:
        cell.font = Font(name='Calibri', size=font_size, bold=font_bold)
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    if border:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    if number_format:
        cell.number_format = number_format
    
    if alignment == 'center':
        cell.alignment = Alignment(horizontal='center', vertical='center')
    elif alignment == 'right':
        cell.alignment = Alignment(horizontal='right', vertical='center')
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center')


def create_constants_sheet(wb):
    """Tworzy arkusz 00_Stałe z parametrami ogólnymi."""
    ws = wb.create_sheet('00_Stałe', 0)
    
    ws['A1'] = 'Parametr'
    ws['B1'] = 'Wartość domyślna'
    ws['C1'] = 'Opis'
    
    for cell in ['A1', 'B1', 'C1']:
        set_cell_style(ws[cell], font_bold=True)
    
    data = [
        ('Min. wkład własny ogółem', 0.20, 'Minimum 20% wartości nieruchomości'),
        ('Min. wkład własny gotówkowy', 0.10, 'Minimum 10% z gotówki (nie z filarów)'),
        ('LTV docelowe po amortyzacji', 0.65, 'Loan-to-Value po spłacie Hypoteki 2'),
        ('Lata amortyzacji do 65%', 15, 'Liczba lat na amortyzację'),
        ('Oprocentowanie testowe (bank)', 0.05, 'Stopa używana przez bank do testu zdolności'),
        ('Roczne koszty utrzymania (test)', 0.01, '1% wartości nieruchomości rocznie'),
        ('Max. Tragbarkeit (udział dochodu)', 0.33, 'Maksymalny udział kosztów w dochodzie'),
        ('Kurs CHF/PLN', 4.60, 'Aktualny kurs franka szwajcarskiego'),
        ('Procent kosztów transakcyjnych (notariusz itd.)', 0.025, 'Szacunkowe koszty notarialne i opłaty'),
    ]
    
    for idx, (param, value, desc) in enumerate(data, start=2):
        ws[f'A{idx}'] = param
        ws[f'B{idx}'] = value
        ws[f'C{idx}'] = desc
        set_cell_style(ws[f'A{idx}'])
        set_cell_style(ws[f'C{idx}'])
        
        if idx in [2, 3, 4, 6, 7, 8, 10]:
            set_cell_style(ws[f'B{idx}'], number_format=FORMAT_PERCENTAGE_00)
        elif idx == 5:
            set_cell_style(ws[f'B{idx}'], number_format='0')
        elif idx == 9:
            set_cell_style(ws[f'B{idx}'], number_format=FORMAT_NUMBER_00)
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50


def create_input_sheet(wb):
    """Tworzy arkusz 01_Wejście z danymi wejściowymi użytkownika."""
    ws = wb.create_sheet('01_Wejście')
    
    ws['A2'] = 'INFORMACJE O NIERUCHOMOŚCI'
    set_cell_style(ws['A2'], font_bold=True, font_size=12, border=False)
    
    ws['A4'] = 'Cena zakupu nieruchomości [CHF]'
    ws['B4'] = ''
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A5'] = 'Szacunkowy % kosztów transakcyjnych'
    ws['B5'] = "='00_Stałe'!B10"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'TWÓJ WKŁAD WŁASNY'
    set_cell_style(ws['A7'], font_bold=True, font_size=12, border=False)
    
    ws['A8'] = 'Wkład gotówkowy [CHF]'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A9'] = 'Wkład z II filaru [CHF]'
    ws['B9'] = ''
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A10'] = 'Wkład z III filaru [CHF]'
    ws['B10'] = ''
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A12'] = 'Suma wkładu własnego [CHF]'
    ws['B12'] = '=B8+B9+B10'
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A13'] = 'Min. wymagany wkład ogółem [CHF]'
    ws['B13'] = "=B4*'00_Stałe'!B2"
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Min. wymagany wkład gotówkowy [CHF]'
    ws['B14'] = "=B4*'00_Stałe'!B3"
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A15'] = 'Status wkładu ogółem'
    ws['B15'] = '=IF(B12>=B13,"OK","ZA MAŁO WKŁADU")'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', font_bold=True)
    
    ws['A16'] = 'Status wkładu gotówkowego'
    ws['B16'] = '=IF(B8>=B14,"OK","ZA MAŁO GOTÓWKI")'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', font_bold=True)
    
    ws['A18'] = 'DANE KREDYTOWE'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A19'] = 'Stopa % Hypoteka 1 (do 65%)'
    ws['B19'] = ''
    set_cell_style(ws['A19'])
    set_cell_style(ws['B19'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A20'] = 'Stopa % Hypoteka 2 (powyżej 65%)'
    ws['B20'] = ''
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A21'] = 'Rodzaj amortyzacji (D – bezpośr.; N – pośrednia)'
    ws['B21'] = ''
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='CCE5FF')
    
    dv = DataValidation(type="list", formula1='"D,N"', allow_blank=True)
    dv.error = 'Wprowadź D lub N'
    dv.errorTitle = 'Nieprawidłowa wartość'
    ws.add_data_validation(dv)
    dv.add(ws['B21'])
    
    ws['A23'] = 'TWOJE FINANSE'
    set_cell_style(ws['A23'], font_bold=True, font_size=12, border=False)
    
    ws['A24'] = 'Dochód brutto gospodarstwa roczny [CHF]'
    ws['B24'] = ''
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A25'] = 'Dobrowolna amortyzacja Hypoteki 1 – rocznie [CHF]'
    ws['B25'] = ''
    set_cell_style(ws['A25'])
    set_cell_style(ws['B25'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A26'] = 'Roczne koszty wspólnoty / Nebenkosten [CHF]'
    ws['B26'] = ''
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A27'] = 'Miesięczny czynsz przy wynajmie porównywalnego lokalu'
    ws['B27'] = ''
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.freeze_panes = 'A3'


def create_financing_sheet(wb):
    """Tworzy arkusz 02_Finansowanie z analizą kredytu."""
    ws = wb.create_sheet('02_Finansowanie')
    
    ws['A1'] = 'PODSTAWY FINANSOWANIA'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Cena zakupu [CHF]'
    ws['B2'] = "='01_Wejście'!B4"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Suma wkładu własnego [CHF]'
    ws['B3'] = "='01_Wejście'!B12"
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A4'] = 'Kwota kredytu (łącznie) [CHF]'
    ws['B4'] = '=B2-B3'
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'LTV początkowe'
    ws['B5'] = '=B4/B2'
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Hypoteka 1 (do 65% wartości) [CHF]'
    ws['B7'] = "=MIN(B4,B2*'00_Stałe'!B4)"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A8'] = 'Hypoteka 2 (powyżej 65%) [CHF]'
    ws['B8'] = '=B4-B7'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A10'] = 'Stopa % Hypoteka 1 (aktualna)'
    ws['B10'] = "='01_Wejście'!B19"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A11'] = 'Stopa % Hypoteka 2 (aktualna)'
    ws['B11'] = "='01_Wejście'!B20"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A13'] = 'Odsetki roczne H1 [CHF]'
    ws['B13'] = '=B7*B10'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Odsetki roczne H2 [CHF]'
    ws['B14'] = '=B8*B11'
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A15'] = 'Razem odsetki roczne [CHF]'
    ws['B15'] = '=B13+B14'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A16'] = 'Odsetki miesięczne [CHF]'
    ws['B16'] = '=B15/12'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A18'] = 'Kwota do amortyzacji (Hypoteka 2) [CHF]'
    ws['B18'] = '=B8'
    set_cell_style(ws['A18'])
    set_cell_style(ws['B18'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A19'] = 'Lata amortyzacji (do 65% LTV)'
    ws['B19'] = "='00_Stałe'!B5"
    set_cell_style(ws['A19'])
    set_cell_style(ws['B19'], bg_color='F2F2F2', font_bold=True, number_format='0')
    
    ws['A20'] = 'Amortyzacja roczna [CHF]'
    ws['B20'] = '=B18/B19'
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A21'] = 'Amortyzacja miesięczna [CHF]'
    ws['B21'] = '=B20/12'
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A22'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B22'] = "='01_Wejście'!B25"
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A23'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B23'] = '=B22/12'
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20


def create_tragbarkeit_sheet(wb):
    """Tworzy arkusz 03_Tragbarkeit z analizą zdolności kredytowej."""
    ws = wb.create_sheet('03_Tragbarkeit')
    
    ws['A1'] = 'TEST ZDOLNOŚCI KREDYTOWEJ (TRAGBARKEIT)'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Kwota kredytu łącznie [CHF]'
    ws['B2'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Stopa testowa banku'
    ws['B3'] = "='00_Stałe'!B6"
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A4'] = 'Odsetki testowe roczne [CHF]'
    ws['B4'] = '=B2*B3'
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'Cena zakupu [CHF]'
    ws['B5'] = "='01_Wejście'!B4"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Koszty utrzymania roczne (test) [CHF]'
    ws['B6'] = "=B5*'00_Stałe'!B7"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A7'] = 'Amortyzacja roczna (test) [CHF]'
    ws['B7'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A9'] = 'Łączne koszty roczne wg banku [CHF]'
    ws['B9'] = '=B4+B6+B7'
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A10'] = 'Dochód brutto roczny [CHF]'
    ws['B10'] = "='01_Wejście'!B24"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A11'] = 'Tragbarkeit (udział dochodu)'
    ws['B11'] = '=B9/B10'
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A12'] = 'Maksymalny udział dochodu'
    ws['B12'] = "='00_Stałe'!B8"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A13'] = 'Ocena Tragbarkeit'
    ws['B13'] = '=IF(B11<=B12,"OK","ZA WYSOKIE OBCIĄŻENIE")'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True)
    
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add('B13', CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill))
    ws.conditional_formatting.add('B13', CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20


def create_cashflow_sheet(wb):
    """Tworzy arkusz 04_Cashflow z analizą rzeczywistych kosztów miesięcznych."""
    ws = wb.create_sheet('04_Cashflow')
    
    ws['A1'] = 'RZECZYWISTE KOSZTY MIESIĘCZNE'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Odsetki miesięczne (aktualne) [CHF]'
    ws['B2'] = "='02_Finansowanie'!B16"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Amortyzacja miesięczna – cash-out [CHF]'
    ws['B3'] = '=IF(\'01_Wejście\'!B21="D",\'02_Finansowanie\'!B21,0)+\'02_Finansowanie\'!B23'
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A4'] = 'Roczne koszty utrzymania realne [CHF]'
    ws['B4'] = "='01_Wejście'!B4*'00_Stałe'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczne koszty utrzymania [CHF]'
    ws['B5'] = '=B4/12'
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Miesięczne koszty wspólnoty (HOA/NK) [CHF]'
    ws['B6'] = "='01_Wejście'!B26/12"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A8'] = 'Łączny miesięczny koszt posiadania [CHF]'
    ws['B8'] = '=B2+B3+B5+B6'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A9'] = 'Kurs CHF/PLN'
    ws['B9'] = "='00_Stałe'!B9"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_NUMBER_00)
    
    ws['A10'] = 'Łączny miesięczny koszt posiadania [PLN]'
    ws['B10'] = '=B8*B9'
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A12'] = 'Miesięczny czynsz przy wynajmie [CHF]'
    ws['B12'] = "='01_Wejście'!B27"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A13'] = 'Różnica: wynajem – posiadanie [CHF/mies.]'
    ws['B13'] = '=B12-B8'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Komentarz'
    ws['B14'] = '=IF(B13>0,"Kupno tańsze od wynajmu","Kupno droższe od wynajmu")'
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True)
    
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add('B14', Rule(type='containsText', operator='containsText', text='tańsze', dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('B14', Rule(type='containsText', operator='containsText', text='droższe', dxf=DifferentialStyle(fill=red_fill)))
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20


def create_yearly_schedule_sheet(wb):
    """Tworzy arkusz 05_Harmonogram_roczny z harmonogramem spłat rocznych."""
    ws = wb.create_sheet('05_Harmonogram_roczny')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'H1 początkowe [CHF]'
    ws['B4'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'H2 początkowe [CHF]'
    ws['B5'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Stopa H1 roczna'
    ws['B6'] = "='01_Wejście'!B19"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Stopa H2 roczna'
    ws['B7'] = "='01_Wejście'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Amortyzacja H2 roczna (obowiązkowa) [CHF]'
    ws['B8'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B9'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A10'] = 'Typ amortyzacji H2 (D/N)'
    ws['B10'] = "='01_Wejście'!B21"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2')
    
    headers = ['Rok', 'Saldo pocz. H1', 'Saldo pocz. H2', 'Saldo pocz. razem', 'Odsetki H1', 'Odsetki H2', 'Odsetki razem',
               'Amortyzacja H2', 'Amortyzacja H1', 'Amortyzacja razem', 'Cash-out roczny', 'Saldo końc. H1', 'Saldo końc. H2', 'Saldo końc. razem']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A13'] = 0
    ws['B13'] = '=$B$4'
    ws['C13'] = '=$B$5'
    ws['D13'] = '=B13+C13'
    ws['E13'] = 0
    ws['F13'] = 0
    ws['G13'] = 0
    ws['H13'] = 0
    ws['I13'] = 0
    ws['J13'] = 0
    ws['K13'] = 0
    ws['L13'] = '=B13'
    ws['M13'] = '=C13'
    ws['N13'] = '=L13+M13'
    
    for col in range(1, 15):
        cell = ws.cell(row=13, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    ws['A14'] = '=A13+1'
    ws['B14'] = '=L13'
    ws['C14'] = '=M13'
    ws['D14'] = '=B14+C14'
    ws['E14'] = '=B14*$B$6'
    ws['F14'] = '=C14*$B$7'
    ws['G14'] = '=E14+F14'
    ws['H14'] = '=IF($B$10="D",MIN($B$8,C14),0)'
    ws['I14'] = '=MIN($B$9,B14)'
    ws['J14'] = '=H14+I14'
    ws['K14'] = '=G14+J14'
    ws['L14'] = '=MAX(0,B14-I14)'
    ws['M14'] = '=MAX(0,C14-H14)'
    ws['N14'] = '=L14+M14'
    
    for col in range(1, 15):
        cell = ws.cell(row=14, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    for row in range(15, 45):
        for col in range(1, 15):
            if col == 1:
                ws.cell(row=row, column=col).value = f'=A{row-1}+1'
            elif col == 2:
                ws.cell(row=row, column=col).value = f'=L{row-1}'
            elif col == 3:
                ws.cell(row=row, column=col).value = f'=M{row-1}'
            elif col == 4:
                ws.cell(row=row, column=col).value = f'=B{row}+C{row}'
            elif col == 5:
                ws.cell(row=row, column=col).value = f'=B{row}*$B$6'
            elif col == 6:
                ws.cell(row=row, column=col).value = f'=C{row}*$B$7'
            elif col == 7:
                ws.cell(row=row, column=col).value = f'=E{row}+F{row}'
            elif col == 8:
                ws.cell(row=row, column=col).value = f'=IF($B$10="D",MIN($B$8,C{row}),0)'
            elif col == 9:
                ws.cell(row=row, column=col).value = f'=MIN($B$9,B{row})'
            elif col == 10:
                ws.cell(row=row, column=col).value = f'=H{row}+I{row}'
            elif col == 11:
                ws.cell(row=row, column=col).value = f'=G{row}+J{row}'
            elif col == 12:
                ws.cell(row=row, column=col).value = f'=MAX(0,B{row}-I{row})'
            elif col == 13:
                ws.cell(row=row, column=col).value = f'=MAX(0,C{row}-H{row})'
            elif col == 14:
                ws.cell(row=row, column=col).value = f'=L{row}+M{row}'
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 14


def create_monthly_schedule_sheet(wb):
    """Tworzy arkusz 06_Harmonogram_miesieczny z harmonogramem spłat miesięcznych."""
    ws = wb.create_sheet('06_Harmonogram_miesieczny')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'H1 początkowe [CHF]'
    ws['B4'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'H2 początkowe [CHF]'
    ws['B5'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Stopa H1 roczna'
    ws['B6'] = "='01_Wejście'!B19"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Stopa H2 roczna'
    ws['B7'] = "='01_Wejście'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Stopa H1 miesięczna'
    ws['B8'] = '=B6/12'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A9'] = 'Stopa H2 miesięczna'
    ws['B9'] = '=B7/12'
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A10'] = 'Amortyzacja H2 roczna [CHF]'
    ws['B10'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A11'] = 'Amortyzacja H2 miesięczna [CHF]'
    ws['B11'] = "='02_Finansowanie'!B21"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A12'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B12'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A13'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B13'] = "='02_Finansowanie'!B23"
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A14'] = 'Typ amortyzacji H2 (D/N)'
    ws['B14'] = "='01_Wejście'!B21"
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2')
    
    headers = ['Miesiąc', 'Saldo pocz. H1', 'Saldo pocz. H2', 'Saldo pocz. razem', 'Odsetki H1', 'Odsetki H2', 'Odsetki razem',
               'Amortyzacja H2', 'Amortyzacja H1', 'Amortyzacja razem', 'Cash-out miesięczny', 'Saldo końc. H1', 'Saldo końc. H2', 'Saldo końc. razem']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=18, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A19'] = 0
    ws['B19'] = '=$B$4'
    ws['C19'] = '=$B$5'
    ws['D19'] = '=B19+C19'
    ws['E19'] = 0
    ws['F19'] = 0
    ws['G19'] = 0
    ws['H19'] = 0
    ws['I19'] = 0
    ws['J19'] = 0
    ws['K19'] = 0
    ws['L19'] = '=B19'
    ws['M19'] = '=C19'
    ws['N19'] = '=L19+M19'
    
    for col in range(1, 15):
        cell = ws.cell(row=19, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    ws['A20'] = '=A19+1'
    ws['B20'] = '=L19'
    ws['C20'] = '=M19'
    ws['D20'] = '=B20+C20'
    ws['E20'] = '=B20*$B$8'
    ws['F20'] = '=C20*$B$9'
    ws['G20'] = '=E20+F20'
    ws['H20'] = '=IF($B$14="D",MIN($B$11,C20),0)'
    ws['I20'] = '=MIN($B$13,B20)'
    ws['J20'] = '=H20+I20'
    ws['K20'] = '=G20+J20'
    ws['L20'] = '=MAX(0,B20-I20)'
    ws['M20'] = '=MAX(0,C20-H20)'
    ws['N20'] = '=L20+M20'
    
    for col in range(1, 15):
        cell = ws.cell(row=20, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    for row in range(21, 381):
        for col in range(1, 15):
            if col == 1:
                ws.cell(row=row, column=col).value = f'=A{row-1}+1'
            elif col == 2:
                ws.cell(row=row, column=col).value = f'=L{row-1}'
            elif col == 3:
                ws.cell(row=row, column=col).value = f'=M{row-1}'
            elif col == 4:
                ws.cell(row=row, column=col).value = f'=B{row}+C{row}'
            elif col == 5:
                ws.cell(row=row, column=col).value = f'=B{row}*$B$8'
            elif col == 6:
                ws.cell(row=row, column=col).value = f'=C{row}*$B$9'
            elif col == 7:
                ws.cell(row=row, column=col).value = f'=E{row}+F{row}'
            elif col == 8:
                ws.cell(row=row, column=col).value = f'=IF($B$14="D",MIN($B$11,C{row}),0)'
            elif col == 9:
                ws.cell(row=row, column=col).value = f'=MIN($B$13,B{row})'
            elif col == 10:
                ws.cell(row=row, column=col).value = f'=H{row}+I{row}'
            elif col == 11:
                ws.cell(row=row, column=col).value = f'=G{row}+J{row}'
            elif col == 12:
                ws.cell(row=row, column=col).value = f'=MAX(0,B{row}-I{row})'
            elif col == 13:
                ws.cell(row=row, column=col).value = f'=MAX(0,C{row}-H{row})'
            elif col == 14:
                ws.cell(row=row, column=col).value = f'=L{row}+M{row}'
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 14


def create_roi_sheet(wb):
    """Tworzy arkusz 07_Analiza_ROI z analizą budowy equity i ROI."""
    ws = wb.create_sheet('07_Analiza_ROI')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wartość zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Wkład własny początkowy [CHF]'
    ws['B5'] = "='01_Wejście'!B12"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Kwota kredytu łącznie [CHF]'
    ws['B6'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Saldo początkowe H1 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Saldo początkowe H2 [CHF]'
    ws['B8'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'Kurs CHF/PLN'
    ws['B9'] = "='00_Stałe'!B9"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format=FORMAT_NUMBER_00)
    
    ws['A10'] = 'Roczny wzrost wartości nieruchomości'
    ws['B10'] = ''
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    headers = ['Rok', 'Wartość nieruchomości [CHF]', 'Saldo H1 [CHF]', 'Saldo H2 [CHF]',
               'Saldo razem [CHF]', 'Equity [CHF]', 'Przyrost equity r/r [CHF]',
               'Skumulowany przyrost equity [CHF]', 'ROI z wkładu (%)']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A16'] = 0
    ws['B16'] = '=$B$4'
    ws['C16'] = '=$B$7'
    ws['D16'] = '=$B$8'
    ws['E16'] = '=C16+D16'
    ws['F16'] = '=B16-E16'
    ws['G16'] = 0
    ws['H16'] = 0
    ws['I16'] = '=F16/$B$5'
    
    for col in range(2, 9):
        cell = ws.cell(row=16, column=col)
        if col <= 8:
            cell.number_format = '#,##0.00'
    ws.cell(row=16, column=9).number_format = FORMAT_PERCENTAGE_00
    
    ws['A17'] = '=A16+1'
    ws['B17'] = '=B16*(1+$B$10)'
    ws['C17'] = "='05_Harmonogram_roczny'!L14"
    ws['D17'] = "='05_Harmonogram_roczny'!M14"
    ws['E17'] = '=C17+D17'
    ws['F17'] = '=B17-E17'
    ws['G17'] = '=F17-F16'
    ws['H17'] = '=H16+G17'
    ws['I17'] = '=F17/$B$5'
    
    for col in range(2, 9):
        cell = ws.cell(row=17, column=col)
        if col <= 8:
            cell.number_format = '#,##0.00'
    ws.cell(row=17, column=9).number_format = FORMAT_PERCENTAGE_00
    
    for row in range(18, 47):
        year = row - 16
        harmonogram_row = 13 + year
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$10)'
        ws.cell(row=row, column=3).value = f"='05_Harmonogram_roczny'!L{harmonogram_row}"
        ws.cell(row=row, column=4).value = f"='05_Harmonogram_roczny'!M{harmonogram_row}"
        ws.cell(row=row, column=5).value = f'=C{row}+D{row}'
        ws.cell(row=row, column=6).value = f'=B{row}-E{row}'
        ws.cell(row=row, column=7).value = f'=F{row}-F{row-1}'
        ws.cell(row=row, column=8).value = f'=H{row-1}+G{row}'
        ws.cell(row=row, column=9).value = f'=F{row}/$B$5'
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            if col <= 8:
                cell.number_format = '#,##0.00'
        ws.cell(row=row, column=9).number_format = FORMAT_PERCENTAGE_00
    
    ws['A50'] = 'Equity po 30 latach [CHF]'
    ws['B50'] = '=F46'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'ROI całkowite z wkładu'
    ws['B51'] = '=I46'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A52'] = 'Średni roczny zwrot CAGR'
    ws['B52'] = '=(1+I46)^(1/30)-1'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 16

# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD create_roi_sheet (przed main)
# ============================================================================

def create_appreciation_sheet(wb):
    """Tworzy arkusz 08_Symulacja_wzrostu_wartości z 3 scenariuszami."""
    ws = wb.create_sheet('08_Symulacja_wzrostu_wartości')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wartość początkowa nieruchomości [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Scenariusz pesymistyczny – roczny wzrost'
    ws['B6'] = ''
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Scenariusz bazowy – roczny wzrost'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Scenariusz optymistyczny – roczny wzrost'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    # Nagłówki tabeli
    headers = ['Rok', 'Pesymistyczny [CHF]', 'Bazowy [CHF]', 'Optymistyczny [CHF]']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0
    ws['A13'] = 0
    ws['B13'] = '=$B$4'
    ws['C13'] = '=$B$4'
    ws['D13'] = '=$B$4'
    
    for col in range(2, 5):
        ws.cell(row=13, column=col).number_format = '#,##0.00'
    
    # Rok 1 (szablon)
    ws['A14'] = '=A13+1'
    ws['B14'] = '=B13*(1+$B$6)'
    ws['C14'] = '=C13*(1+$B$7)'
    ws['D14'] = '=D13*(1+$B$8)'
    
    for col in range(2, 5):
        ws.cell(row=14, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 43)
    for row in range(15, 44):
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$6)'
        ws.cell(row=row, column=3).value = f'=C{row-1}*(1+$B$7)'
        ws.cell(row=row, column=4).value = f'=D{row-1}*(1+$B$8)'
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A46'] = 'Wartość po 30 latach – pesymistyczny'
    ws['B46'] = '=B43'
    set_cell_style(ws['A46'], font_bold=True)
    set_cell_style(ws['B46'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A47'] = 'Wartość po 30 latach – bazowy'
    ws['B47'] = '=C43'
    set_cell_style(ws['A47'], font_bold=True)
    set_cell_style(ws['B47'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A48'] = 'Wartość po 30 latach – optymistyczny'
    ws['B48'] = '=D43'
    set_cell_style(ws['A48'], font_bold=True)
    set_cell_style(ws['B48'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A50'] = 'CAGR scenariusz bazowy'
    ws['B50'] = '=(C43/$B$4)^(1/30)-1'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_appreciation_sheet (przed main)
# ============================================================================

def create_opportunity_cost_sheet(wb):
    """Tworzy arkusz 09_Koszt_alternatywny_kapitalu - porównanie ETF vs equity."""
    ws = wb.create_sheet('09_Koszt_alternatywny_kapitalu')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wkład własny początkowy [CHF]'
    ws['B4'] = "='01_Wejście'!B12"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Roczna stopa zwrotu alternatywnej inwestycji'
    ws['B5'] = ''
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A6'] = 'Roczna kwota odpowiadająca amortyzacji H2 [CHF]'
    ws['B6'] = '=IF(\'01_Wejście\'!B21="D",\'02_Finansowanie\'!B20,0)'
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Roczna kwota odpowiadająca dobrowolnej amortyzacji H1 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Łączna roczna kwota dodatkowych wpłat [CHF]'
    ws['B8'] = '=B6+B7'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli
    headers = ['Rok', 'Kapitał pocz. w roku [CHF]', 'Nowe wpłaty w roku [CHF]', 
               'Zysk z inwestycji w roku [CHF]', 'Kapitał końcowy [CHF]', 
               'Equity w nieruchomości [CHF]', 'Różnica: ETF – equity [CHF]', 'Komentarz']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz 16)
    ws['A16'] = 0
    ws['B16'] = '=$B$4'
    ws['C16'] = 0
    ws['D16'] = 0
    ws['E16'] = '=B16+C16+D16'
    ws['F16'] = "='07_Analiza_ROI'!F16"
    ws['G16'] = '=E16-F16'
    ws['H16'] = '=IF(G16>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
    
    for col in range(2, 8):
        ws.cell(row=16, column=col).number_format = '#,##0.00'
    
    # Rok 1 (wiersz 17 - szablon)
    ws['A17'] = '=A16+1'
    ws['B17'] = '=E16'
    ws['C17'] = '=$B$8'
    ws['D17'] = '=(B17+C17)*$B$5'
    ws['E17'] = '=B17+C17+D17'
    ws['F17'] = "='07_Analiza_ROI'!F17"
    ws['G17'] = '=E17-F17'
    ws['H17'] = '=IF(G17>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
    
    for col in range(2, 8):
        ws.cell(row=17, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 46)
    for row in range(18, 47):
        year = row - 16
        roi_row = 16 + year
        
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=E{row-1}'
        ws.cell(row=row, column=3).value = '=$B$8'
        ws.cell(row=row, column=4).value = f'=(B{row}+C{row})*$B$5'
        ws.cell(row=row, column=5).value = f'=B{row}+C{row}+D{row}'
        ws.cell(row=row, column=6).value = f"='07_Analiza_ROI'!F{roi_row}"
        ws.cell(row=row, column=7).value = f'=E{row}-F{row}'
        ws.cell(row=row, column=8).value = f'=IF(G{row}>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
        
        for col in range(2, 8):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A50'] = 'Kapitał alternatywny po 30 latach [CHF]'
    ws['B50'] = '=E46'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'Equity w nieruchomości po 30 latach [CHF]'
    ws['B51'] = '=F46'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A52'] = 'Różnica: ETF – nieruchomość [CHF]'
    ws['B52'] = '=B50-B51'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A53'] = 'Komentarz'
    ws['B53'] = '=IF(B52>0,"Lepsza inwestycja w ETF","Lepsza inwestycja w nieruchomość")'
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['H'].width = 25



# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD definicją create_opportunity_cost_sheet (lub ostatnią funkcją tworzącą arkusze)
# ============================================================================

def create_rent_vs_buy_sheet(wb):
    """Tworzy arkusz 10_Rent_vs_Buy_30lat - porównanie kupna vs wynajmu w 30 latach."""
    ws = wb.create_sheet('10_Rent_vs_Buy_30lat')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Miesięczny koszt posiadania [CHF]'
    ws['B4'] = "='04_Cashflow'!B8"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczny czynsz wynajmu [CHF]'
    ws['B5'] = "='01_Wejście'!B27"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Roczny wzrost czynszu'
    ws['B6'] = ''
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Roczny wzrost kosztów posiadania'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Equity po 30 latach [CHF]'
    ws['B8'] = "='07_Analiza_ROI'!F46"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli porównania
    headers = ['Rok', 'Koszt posiadania (Kupno) – roczny [CHF]', 'Koszt wynajmu – roczny [CHF]',
               'Różnica (Wynajem – Kupno) [CHF]', 'Skumulowana różnica [CHF]']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz 16)
    ws['A16'] = 0
    ws['B16'] = '=$B$4*12'
    ws['C16'] = '=$B$5*12'
    ws['D16'] = '=C16-B16'
    ws['E16'] = '=D16'
    
    for col in range(2, 6):
        ws.cell(row=16, column=col).number_format = '#,##0.00'
    
    # Rok 1 (wiersz 17 - szablon)
    ws['A17'] = '=A16+1'
    ws['B17'] = '=B16*(1+$B$7)'
    ws['C17'] = '=C16*(1+$B$6)'
    ws['D17'] = '=C17-B17'
    ws['E17'] = '=E16+D17'
    
    for col in range(2, 6):
        ws.cell(row=17, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 46)
    for row in range(18, 47):
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$7)'
        ws.cell(row=row, column=3).value = f'=C{row-1}*(1+$B$6)'
        ws.cell(row=row, column=4).value = f'=C{row}-B{row}'
        ws.cell(row=row, column=5).value = f'=E{row-1}+D{row}'
        
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A50'] = 'Suma kosztów posiadania przez 30 lat [CHF]'
    ws['B50'] = '=SUM(B16:B46)'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'Suma kosztów wynajmu przez 30 lat [CHF]'
    ws['B51'] = '=SUM(C16:C46)'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A52'] = 'Różnica (wynajem – kupno) [CHF]'
    ws['B52'] = '=B51-B50'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A53'] = 'Equity po 30 latach [CHF]'
    ws['B53'] = '=$B$8'
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A54'] = 'Efekt netto (różnica + equity) [CHF]'
    ws['B54'] = '=B52+B53'
    set_cell_style(ws['A54'], font_bold=True)
    set_cell_style(ws['B54'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A55'] = 'Komentarz'
    ws['B55'] = '=IF(B54>0,"Kupno opłaca się bardziej","Wynajem opłaca się bardziej")'
    set_cell_style(ws['A55'], font_bold=True)
    set_cell_style(ws['B55'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_rent_vs_buy_sheet
# ============================================================================

def create_stress_test_sheet(wb):
    """Tworzy arkusz 11_Stress_test - szok stóp procentowych a koszty i Tragbarkeit."""
    ws = wb.create_sheet('11_Stress_test')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Stopa H1 aktualna'
    ws['B4'] = "='01_Wejście'!B19"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A5'] = 'Stopa H2 aktualna'
    ws['B5'] = "='01_Wejście'!B20"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A6'] = 'Saldo H1 [CHF]'
    ws['B6'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Saldo H2 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Koszt utrzymania miesięczny [CHF]'
    ws['B8'] = "='04_Cashflow'!B5"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'HOA/Nebenkosten miesięczne [CHF]'
    ws['B9'] = "='04_Cashflow'!B6"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A10'] = 'Amortyzacja H2 miesięczna [CHF]'
    ws['B10'] = "='02_Finansowanie'!B21"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A11'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B11'] = "='02_Finansowanie'!B23"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A12'] = 'Miesięczny czynsz wynajmu [CHF]'
    ws['B12'] = "='01_Wejście'!B27"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli scenariuszy
    headers = ['Szok stopy (Δ, w p.p.)', 'Odsetki H1 miesięczne [CHF]', 'Odsetki H2 miesięczne [CHF]',
               'Łączne odsetki miesięczne [CHF]', 'Miesięczny cash-out (posiadanie) [CHF]',
               'Różnica vs wynajem [CHF]', 'Tragbarkeit (udział dochodu)', 'Status']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Scenariusze szoku
    shocks = [0, 0.005, 0.01, 0.015, 0.02]
    
    for idx, shock in enumerate(shocks):
        row = 16 + idx
        
        # Szok w kolumnie A
        ws.cell(row=row, column=1).value = shock
        ws.cell(row=row, column=1).number_format = FORMAT_PERCENTAGE_00
        
        # B: Odsetki H1 miesięczne
        ws.cell(row=row, column=2).value = f'=$B$6*(($B$4+A{row})/12)'
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        # C: Odsetki H2 miesięczne
        ws.cell(row=row, column=3).value = f'=$B$7*(($B$5+A{row})/12)'
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # D: Łączne odsetki
        ws.cell(row=row, column=4).value = f'=B{row}+C{row}'
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        
        # E: Miesięczny cash-out
        ws.cell(row=row, column=5).value = f'=D{row}+$B$8+$B$9+$B$10+$B$11'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        
        # F: Różnica vs wynajem
        ws.cell(row=row, column=6).value = f'=$B$12-E{row}'
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        # G: Tragbarkeit
        ws.cell(row=row, column=7).value = f"=(E{row}*12)/'01_Wejście'!B24"
        ws.cell(row=row, column=7).number_format = FORMAT_PERCENTAGE_00
        
        # H: Status
        ws.cell(row=row, column=8).value = f"=IF(G{row}<'00_Stałe'!B8,\"OK\",\"Ryzyko Tragbarkeit\")"
    
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 22
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20


# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD create_stress_test_sheet (lub ostatnią funkcją tworzącą arkusze)
# ============================================================================

def create_sale_analysis_sheet(wb):
    """Tworzy arkusz 12_Analiza_sprzedaży_po_X_latach - analiza wyniku sprzedaży po X latach."""
    ws = wb.create_sheet('12_Analiza_sprzedazy_X_lat')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Cena zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Wkład własny początkowy [CHF]'
    ws['B5'] = "='01_Wejście'!B12"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Kwota kredytu łącznie [CHF]'
    ws['B6'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Lata do sprzedaży (X)'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A8'] = 'Roczny wzrost wartości nieruchomości'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A9'] = 'Prowizja/koszty sprzedaży [% od ceny]'
    ws['B9'] = ''
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A10'] = 'Rok sprzedaży w harmonogramie'
    ws['B10'] = '=13+$B$7'
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='0')
    
    # Wycena nieruchomości przy sprzedaży
    ws['A13'] = 'WYCENA NIERUCHOMOŚCI PRZY SPRZEDAŻY'
    set_cell_style(ws['A13'], font_bold=True, font_size=12, border=False)
    
    ws['A15'] = 'Wartość nieruchomości przy sprzedaży [CHF]'
    ws['B15'] = '=B4*(1+$B$8)^$B$7'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A16'] = 'Koszty sprzedaży (prowizje, podatki) [CHF]'
    ws['B16'] = '=B15*$B$9'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Saldo kredytu przy sprzedaży
    ws['A18'] = 'SALDO KREDYTU PRZY SPRZEDAŻY'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Saldo końcowe H1 w roku sprzedaży [CHF]'
    ws['B20'] = "=INDEX('05_Harmonogram_roczny'!L:L,$B$10)"
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A21'] = 'Saldo końcowe H2 w roku sprzedaży [CHF]'
    ws['B21'] = "=INDEX('05_Harmonogram_roczny'!M:M,$B$10)"
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A22'] = 'Saldo kredytu łącznie [CHF]'
    ws['B22'] = '=B20+B21'
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Rozliczenie sprzedaży
    ws['A24'] = 'ROZLICZENIE SPRZEDAŻY'
    set_cell_style(ws['A24'], font_bold=True, font_size=12, border=False)
    
    ws['A26'] = 'Cena sprzedaży brutto [CHF]'
    ws['B26'] = '=B15'
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A27'] = 'Koszty sprzedaży [CHF]'
    ws['B27'] = '=B16'
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A28'] = 'Przychód netto po kosztach sprzedaży [CHF]'
    ws['B28'] = '=B26-B27'
    set_cell_style(ws['A28'])
    set_cell_style(ws['B28'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A29'] = 'Spłata kredytu (H1 + H2) [CHF]'
    ws['B29'] = '=B22'
    set_cell_style(ws['A29'])
    set_cell_style(ws['B29'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A30'] = 'Środki po spłacie kredytu [CHF]'
    ws['B30'] = '=B28-B29'
    set_cell_style(ws['A30'], font_bold=True)
    set_cell_style(ws['B30'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    # Analiza zwrotu
    ws['A32'] = 'ANALIZA ZWROTU Z INWESTYCJI'
    set_cell_style(ws['A32'], font_bold=True, font_size=12, border=False)
    
    ws['A34'] = 'Zysk brutto względem wkładu własnego [CHF]'
    ws['B34'] = '=B30-B5'
    set_cell_style(ws['A34'], font_bold=True)
    set_cell_style(ws['B34'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A35'] = 'ROI względem wkładu własnego'
    ws['B35'] = '=B34/B5'
    set_cell_style(ws['A35'], font_bold=True)
    set_cell_style(ws['B35'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A36'] = 'Średni roczny zwrot (CAGR)'
    ws['B36'] = '=(1+B35)^(1/$B$7)-1'
    set_cell_style(ws['A36'], font_bold=True)
    set_cell_style(ws['B36'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A38'] = 'Komentarz'
    ws['B38'] = '=IF(B35>0,"Zysk na inwestycji","Strata na inwestycji")'
    set_cell_style(ws['A38'], font_bold=True)
    set_cell_style(ws['B38'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_sale_analysis_sheet
# ============================================================================

def create_prd_analysis_sheet(wb):
    """Tworzy arkusz 13_Analiza_PRD - Price-to-Rent Ratio, yield i interpretacja wyceny."""
    ws = wb.create_sheet('13_Analiza_PRD')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Cena zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczny czynsz referencyjny [CHF]'
    ws['B5'] = "='01_Wejście'!B27"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Roczny czynsz referencyjny [CHF]'
    ws['B6'] = '=B5*12'
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Cena zakupu + koszty transakcyjne [CHF]'
    ws['B7'] = "=B4*(1+'00_Stałe'!B10)"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Wskaźniki PRD i yield
    ws['A10'] = 'Wskaźniki wyceny (Price-to-Rent Ratio)'
    set_cell_style(ws['A10'], font_bold=True, font_size=12, border=False)
    
    ws['A12'] = 'Price-to-Rent Ratio (PRD) – bez kosztów'
    ws['B12'] = '=B4/B6'
    set_cell_style(ws['A12'], font_bold=True)
    set_cell_style(ws['B12'], font_bold=True, bg_color='FFEB9C', number_format='0.0')
    
    ws['A13'] = 'Price-to-Rent (z kosztami transakcyjnymi)'
    ws['B13'] = '=B7/B6'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', number_format='0.0')
    
    ws['A14'] = 'Brutto yield (roczny czynsz / cena)'
    ws['B14'] = '=B6/B4'
    set_cell_style(ws['A14'], font_bold=True)
    set_cell_style(ws['B14'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A15'] = 'Brutto yield (z kosztami)'
    ws['B15'] = '=B6/B7'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Interpretacja PRD
    ws['A18'] = 'Interpretacja wskaźnika Price-to-Rent'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Klasyfikacja (na podstawie PRD)'
    ws['B20'] = '=IF(B12<15,"Tanio",IF(B12<20,"Normalnie",IF(B12<25,"Drogo","Bardzo drogo")))'
    set_cell_style(ws['A20'], font_bold=True)
    set_cell_style(ws['B20'], font_bold=True, bg_color='FFEB9C')
    
    # Dodatkowe formatowanie warunkowe dla B20
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Tanio', dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Normalnie', dxf=DifferentialStyle(fill=yellow_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Drogo', dxf=DifferentialStyle(fill=orange_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Bardzo drogo', dxf=DifferentialStyle(fill=red_fill)))
    
    # Porównanie z kosztem posiadania
    ws['A23'] = 'Koszt posiadania miesięcznie [CHF]'
    ws['B23'] = "='04_Cashflow'!B8"
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A24'] = 'Różnica: czynsz – koszt posiadania [CHF/mies.]'
    ws['B24'] = '=B5-B23'
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30

def create_new_property_after_sale_sheet(wb):
    """Tworzy arkusz 14_Nowa_nieruchomosc_X_lat - analiza maksymalnej ceny nowej nieruchomości po sprzedaży."""
    ws = wb.create_sheet('14_Nowa_nieruchomosc_X_lat')
    
    # Nagłówek
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    # ========================================================================
    # SEKCJA A – Dane z innych arkuszy (tylko odczyt, szare tło)
    # ========================================================================
    
    ws['A4'] = 'DANE Z OBECNEJ NIERUCHOMOŚCI I SPRZEDAŻY'
    set_cell_style(ws['A4'], font_bold=True, font_size=12, border=False)
    
    ws['A6'] = 'Środki po sprzedaży (po spłacie kredytu) [CHF]'
    ws['B6'] = "='12_Analiza_sprzedaży_po_X_latach'!B30"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A7'] = 'Dochód brutto roczny dziś [CHF]'
    ws['B7'] = "='01_Wejście'!B24"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A8'] = 'Lata do sprzedaży obecnej nieruchomości (X)'
    ws['B8'] = "='12_Analiza_sprzedaży_po_X_latach'!B7"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', font_bold=True, number_format='0')
    
    ws['A10'] = 'PARAMETRY BANKOWE (z arkusza 00_Stałe)'
    set_cell_style(ws['A10'], font_bold=True, font_size=12, border=False)
    
    ws['A12'] = 'Min. wkład własny ogółem [%]'
    ws['B12'] = "='00_Stałe'!B2"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A13'] = 'LTV docelowe po amortyzacji'
    ws['B13'] = "='00_Stałe'!B4"
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A14'] = 'Lata amortyzacji do docelowego LTV'
    ws['B14'] = "='00_Stałe'!B5"
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True, number_format='0')
    
    ws['A15'] = 'Stopa testowa banku'
    ws['B15'] = "='00_Stałe'!B6"
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A16'] = 'Roczne koszty utrzymania (test) [% od ceny]'
    ws['B16'] = "='00_Stałe'!B7"
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A17'] = 'Maksymalny udział dochodu (Tragbarkeit)'
    ws['B17'] = "='00_Stałe'!B8"
    set_cell_style(ws['A17'])
    set_cell_style(ws['B17'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA B – Parametry użytkownika dla nowej nieruchomości (niebieskie pola)
    # ========================================================================
    
    ws['A20'] = 'PARAMETRY SCENARIUSZA ZAKUPU NOWEJ NIERUCHOMOŚCI'
    set_cell_style(ws['A20'], font_bold=True, font_size=12, border=False)
    
    ws['A22'] = 'Roczny wzrost dochodu [%]'
    ws['B22'] = ''
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A23'] = 'Dodatkowe roczne oszczędności [CHF]'
    ws['B23'] = ''
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A24'] = 'Część środków ze sprzedaży przeznaczona na wkład [%]'
    ws['B24'] = 1.0
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA C – Obliczona przyszła pozycja finansowa
    # ========================================================================
    
    ws['A27'] = 'OBLICZONA PRZYSZŁA POZYCJA FINANSOWA'
    set_cell_style(ws['A27'], font_bold=True, font_size=12, border=False)
    
    ws['A29'] = 'Dochód brutto w roku zakupu nowej nieruchomości [CHF]'
    ws['B29'] = '=B7*(1+B22)^B8'
    set_cell_style(ws['A29'])
    set_cell_style(ws['B29'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A30'] = 'Skumulowane dodatkowe oszczędności do momentu sprzedaży [CHF]'
    ws['B30'] = '=B23*B8'
    set_cell_style(ws['A30'])
    set_cell_style(ws['B30'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A31'] = 'Środki ze sprzedaży przeznaczone na wkład [CHF]'
    ws['B31'] = '=B6*B24'
    set_cell_style(ws['A31'])
    set_cell_style(ws['B31'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A32'] = 'Łączny dostępny wkład własny na nową nieruchomość [CHF]'
    ws['B32'] = '=B30+B31'
    set_cell_style(ws['A32'], font_bold=True)
    set_cell_style(ws['B32'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA D – Test banku dla nowej nieruchomości
    # ========================================================================
    
    ws['A35'] = 'TEST BANKU DLA NOWEJ NIERUCHOMOŚCI'
    set_cell_style(ws['A35'], font_bold=True, font_size=12, border=False)
    
    ws['A37'] = 'Testowa cena nowej nieruchomości [CHF]'
    ws['B37'] = ''
    set_cell_style(ws['A37'])
    set_cell_style(ws['B37'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A38'] = 'Kwota kredytu nowej nieruchomości [CHF]'
    ws['B38'] = '=B37-B32'
    set_cell_style(ws['A38'])
    set_cell_style(ws['B38'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A39'] = 'LTV początkowe'
    ws['B39'] = '=IF(B37>0,B38/B37,0)'
    set_cell_style(ws['A39'])
    set_cell_style(ws['B39'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A40'] = 'Kwota do amortyzacji (H2 nowej nieruchomości) [CHF]'
    ws['B40'] = '=IF(B37>0,MAX(0,B38-B37*B13),0)'
    set_cell_style(ws['A40'])
    set_cell_style(ws['B40'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A41'] = 'Amortyzacja roczna [CHF]'
    ws['B41'] = '=B40/B14'
    set_cell_style(ws['A41'])
    set_cell_style(ws['B41'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A42'] = 'Odsetki testowe roczne [CHF]'
    ws['B42'] = '=B38*B15'
    set_cell_style(ws['A42'])
    set_cell_style(ws['B42'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A43'] = 'Koszty utrzymania roczne [CHF]'
    ws['B43'] = '=B37*B16'
    set_cell_style(ws['A43'])
    set_cell_style(ws['B43'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A44'] = 'Łączne koszty roczne wg banku [CHF]'
    ws['B44'] = '=B41+B42+B43'
    set_cell_style(ws['A44'])
    set_cell_style(ws['B44'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A45'] = 'Tragbarkeit przy nowej nieruchomości'
    ws['B45'] = '=IF(B29>0,B44/B29,0)'
    set_cell_style(ws['A45'], font_bold=True)
    set_cell_style(ws['B45'], bg_color='FFEB9C', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A46'] = 'Limit Tragbarkeit'
    ws['B46'] = '=B17'
    set_cell_style(ws['A46'])
    set_cell_style(ws['B46'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A47'] = 'Ocena Tragbarkeit nowej nieruchomości'
    ws['B47'] = '=IF(B45<=B46,"OK","ZA WYSOKIE OBCIĄŻENIE")'
    set_cell_style(ws['A47'], font_bold=True)
    set_cell_style(ws['B47'], bg_color='F2F2F2', font_bold=True)
    
    ws['A49'] = 'Min. wymagany wkład własny [CHF]'
    ws['B49'] = '=B37*B12'
    set_cell_style(ws['A49'])
    set_cell_style(ws['B49'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A50'] = 'Ocena wkładu własnego'
    ws['B50'] = '=IF(B32>=B49,"OK","ZA MAŁO WKŁADU")'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], bg_color='F2F2F2', font_bold=True)
    
    # Formatowanie warunkowe dla ocen
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    ws.conditional_formatting.add('B47', CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill))
    ws.conditional_formatting.add('B47', CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
    
    ws.conditional_formatting.add('B50', CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill))
    ws.conditional_formatting.add('B50', CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
    
    # ========================================================================
    # SEKCJA E (opcjonalna) – Tabela scenariuszy dla różnych cen
    # ========================================================================
    
    ws['A53'] = 'TABELA SCENARIUSZY – RÓŻNE CENY NOWEJ NIERUCHOMOŚCI'
    set_cell_style(ws['A53'], font_bold=True, font_size=12, border=False)
    
    headers_scenario = ['Cena nowej nieruchomości [CHF]', 'LTV początkowe', 
                        'Tragbarkeit', 'Status Tragbarkeit', 'Status wkładu']
    
    for col_idx, header in enumerate(headers_scenario, start=1):
        cell = ws.cell(row=55, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Przykładowe scenariusze – 10 wariantów ceny (użytkownik może edytować kolumnę A)
    for row_offset in range(10):
        row = 56 + row_offset
        
        # A: Cena (puste – użytkownik może wpisać)
        ws.cell(row=row, column=1).value = ''
        set_cell_style(ws.cell(row=row, column=1), bg_color='CCE5FF', number_format='#,##0.00')
        
        # B: LTV
        ws.cell(row=row, column=2).value = f'=IF(A{row}>0,(A{row}-$B$32)/A{row},0)'
        ws.cell(row=row, column=2).number_format = FORMAT_PERCENTAGE_00
        
        # C: Tragbarkeit
        ws.cell(row=row, column=3).value = f'=IF(AND(A{row}>0,$B$29>0),((MAX(0,(A{row}-$B$32)-A{row}*$B$13)/$B$14)+(A{row}-$B$32)*$B$15+A{row}*$B$16)/$B$29,0)'
        ws.cell(row=row, column=3).number_format = FORMAT_PERCENTAGE_00
        
        # D: Status Tragbarkeit
        ws.cell(row=row, column=4).value = f'=IF(A{row}="","",IF(C{row}<=$B$17,"OK","Ryzyko"))'
        
        # E: Status wkładu
        ws.cell(row=row, column=5).value = f'=IF(A{row}="","",IF($B$32>=A{row}*$B$12,"OK","Za mało"))'
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
# ============================================================================
# 3) ZAMIENIĆ dotychczasową funkcję main() na poniższą wersję
# ============================================================================

def create_family_planning_sheet(wb):
    """Tworzy arkusz 15_Planowanie_rodziny - symulacja zmian dochodu po narodzinach dzieci."""
    ws = wb.create_sheet('15_Planowanie_rodziny')
    
    # Nagłówek główny
    ws['A1'] = 'PLANOWANIE RODZINY I WPŁYW NA DOCHÓD'
    set_cell_style(ws['A1'], font_bold=True, font_size=14, border=False)
    
    # ========================================================================
    # SEKCJA A – Dane wejściowe (rodzina)
    # ========================================================================
    
    ws['A3'] = 'DANE RODZINY I DZIECI'
    set_cell_style(ws['A3'], font_bold=True, font_size=12, border=False)
    
    ws['A5'] = 'Liczba dzieci'
    ws['B5'] = ''
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='CCE5FF', number_format='0')
    
    ws['A6'] = 'Rok pojawienia się pierwszego dziecka'
    ws['B6'] = ''
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format='0')
    
    ws['A7'] = 'Rok pojawienia się drugiego dziecka (jeśli B5 ≥ 2)'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A9'] = 'ZMIANA ETATÓW RODZICÓW'
    set_cell_style(ws['A9'], font_bold=True, font_size=12, border=False)
    
    ws['A11'] = 'Zmiana etatu partnera 1 (1.0 = 100%)'
    ws['B11'] = 1.0
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A12'] = 'Zmiana etatu partnera 2 (1.0 = 100%)'
    ws['B12'] = 1.0
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A14'] = 'KOSZTY OPIEKI NAD DZIEĆMI'
    set_cell_style(ws['A14'], font_bold=True, font_size=12, border=False)
    
    ws['A16'] = 'Koszt Kita/Tagesfamilie na dziecko [CHF/mies]'
    ws['B16'] = ''
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A17'] = 'Liczba dni opieki w tygodniu (0–5)'
    ws['B17'] = ''
    set_cell_style(ws['A17'])
    set_cell_style(ws['B17'], bg_color='CCE5FF', number_format='0.0')
    
    ws['A18'] = 'Subsydia gminne na dziecko [CHF/mies]'
    ws['B18'] = ''
    set_cell_style(ws['A18'])
    set_cell_style(ws['B18'], bg_color='CCE5FF', number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA B – Kinderzulage i podatki
    # ========================================================================
    
    ws['A21'] = 'KINDERZULAGE I PODATKI'
    set_cell_style(ws['A21'], font_bold=True, font_size=12, border=False)
    
    ws['A23'] = 'Kinderzulage podstawowa [CHF/mies]'
    ws['B23'] = 250
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A24'] = 'Dodatkowa Kinderzulage (Ausbildungszulage) [CHF/mies]'
    ws['B24'] = ''
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A26'] = 'Dochód brutto roczny przed dziećmi [CHF]'
    ws['B26'] = "='01_Wejście'!B24"
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A27'] = 'Zmiana podatku po dzieciach [CHF/mies]'
    ws['B27'] = ''
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A28'] = 'Opis zmiany podatku'
    ws['B28'] = '(ujemne = ulga podatkowa)'
    set_cell_style(ws['A28'])
    set_cell_style(ws['B28'], bg_color='F2F2F2')
    
    # ========================================================================
    # SEKCJA C – Obliczenia rdzeniowe
    # ========================================================================
    
    ws['A31'] = 'OBLICZENIA – WPŁYW NA DOCHÓD'
    set_cell_style(ws['A31'], font_bold=True, font_size=12, border=False)
    
    ws['A33'] = 'Dochód roczny teraz [CHF]'
    ws['B33'] = '=B26'
    set_cell_style(ws['A33'])
    set_cell_style(ws['B33'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A34'] = 'Średni etat partnera (założenie uproszczone)'
    ws['B34'] = '=(B11+B12)/2'
    set_cell_style(ws['A34'])
    set_cell_style(ws['B34'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A35'] = 'Dochód po zmianie etatu [CHF/rok]'
    ws['B35'] = '=B33*B34'
    set_cell_style(ws['A35'])
    set_cell_style(ws['B35'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A37'] = 'Kinderzulage roczna (suma) [CHF]'
    ws['B37'] = '=(B23+B24)*B5*12'
    set_cell_style(ws['A37'])
    set_cell_style(ws['B37'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A39'] = 'Koszt opieki miesięczny [CHF]'
    ws['B39'] = '=IF(B5>0,B5*B16*(B17/5)-B5*B18,0)'
    set_cell_style(ws['A39'])
    set_cell_style(ws['B39'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A40'] = 'Koszt opieki roczny [CHF]'
    ws['B40'] = '=B39*12'
    set_cell_style(ws['A40'])
    set_cell_style(ws['B40'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A42'] = 'Zmiana podatku roczna [CHF]'
    ws['B42'] = '=B27*12'
    set_cell_style(ws['A42'])
    set_cell_style(ws['B42'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A44'] = 'Dochód netto po zmianach [CHF/rok]'
    ws['B44'] = '=B35-B40+B37-B42'
    set_cell_style(ws['A44'], font_bold=True)
    set_cell_style(ws['B44'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A45'] = 'Zmiana dochodu netto [CHF/rok]'
    ws['B45'] = '=B44-B33'
    set_cell_style(ws['A45'], font_bold=True)
    set_cell_style(ws['B45'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A46'] = 'Zmiana dochodu netto [%]'
    ws['B46'] = '=IF(B33>0,B45/B33,0)'
    set_cell_style(ws['A46'], font_bold=True)
    set_cell_style(ws['B46'], bg_color='FFEB9C', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA D – Eksport danych do innych arkuszy
    # ========================================================================
    
    ws['A50'] = 'EKSPORT DO INNYCH ARKUSZY'
    set_cell_style(ws['A50'], font_bold=True, font_size=12, border=False)
    
    ws['A52'] = 'Dochód skorygowany po planowaniu rodziny [CHF/rok]'
    ws['B52'] = '=B44'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A53'] = 'Dochód skorygowany [CHF/mies]'
    ws['B53'] = '=B44/12'
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A55'] = 'Użyj dochodu skorygowanego w modelu? (0/1)'
    ws['B55'] = 0
    set_cell_style(ws['A55'], font_bold=True)
    set_cell_style(ws['B55'], bg_color='CCE5FF', number_format='0')
    
    ws['A57'] = 'Dochód przekazywany do innych arkuszy [CHF/rok]'
    ws['B57'] = "=IF(B55=1,B52,'01_Wejście'!B24)"
    set_cell_style(ws['A57'], font_bold=True)
    set_cell_style(ws['B57'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A59'] = 'UWAGA: Komórka B57 może być użyta w arkuszach:'
    ws['B59'] = '03_Tragbarkeit, 11_Stress_test, 14_Nowa_nieruchomosc_X_lat'
    set_cell_style(ws['A59'])
    set_cell_style(ws['B59'], bg_color='FFF2CC')
    
    ws['A60'] = 'Aby aktywować:'
    ws['B60'] = 'Ustaw B55 = 1 i zamień odniesienia do dochodu na =15_Planowanie_rodziny!B57'
    set_cell_style(ws['A60'])
    set_cell_style(ws['B60'], bg_color='FFF2CC')
    
    # ========================================================================
    # SEKCJA E (opcjonalna) – Podsumowanie w czasie
    # ========================================================================
    
    ws['A64'] = 'PODSUMOWANIE MIESIĘCZNE'
    set_cell_style(ws['A64'], font_bold=True, font_size=12, border=False)
    
    ws['A66'] = 'Dochód miesięczny przed zmianami [CHF]'
    ws['B66'] = '=B33/12'
    set_cell_style(ws['A66'])
    set_cell_style(ws['B66'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A67'] = 'Dochód miesięczny po zmianach [CHF]'
    ws['B67'] = '=B44/12'
    set_cell_style(ws['A67'])
    set_cell_style(ws['B67'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A68'] = 'Kinderzulage miesięcznie [CHF]'
    ws['B68'] = '=B37/12'
    set_cell_style(ws['A68'])
    set_cell_style(ws['B68'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A69'] = 'Koszty opieki miesięcznie [CHF]'
    ws['B69'] = '=B39'
    set_cell_style(ws['A69'])
    set_cell_style(ws['B69'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A70'] = 'Zmiana podatku miesięcznie [CHF]'
    ws['B70'] = '=B27'
    set_cell_style(ws['A70'])
    set_cell_style(ws['B70'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A71'] = 'Różnica miesięczna netto [CHF]'
    ws['B71'] = '=B67-B66'
    set_cell_style(ws['A71'], font_bold=True)
    set_cell_style(ws['B71'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Formatowanie warunkowe dla B71
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    ws.conditional_formatting.add('B71', CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    ws.conditional_formatting.add('B71', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 30


def create_tax_canton_analysis_sheet(wb):
    """Tworzy arkusz 17_Podatki_kantony - analiza podatkowa kantonów."""
    ws = wb.create_sheet('17_Podatki_kantony')
    
    # Nagłówek główny
    ws['A1'] = 'ANALIZA PODATKOWA – PORÓWNANIE KANTONÓW'
    set_cell_style(ws['A1'], font_bold=True, font_size=14, border=False)
    
    # ========================================================================
    # SEKCJA A – Wybór dochodu (źródło)
    # ========================================================================
    
    ws['A3'] = 'DANE DOCHODOWE (ŹRÓDŁO)'
    set_cell_style(ws['A3'], font_bold=True, font_size=12, border=False)
    
    ws['A5'] = 'Dochód brutto roczny (bazowy) [CHF]'
    ws['B5'] = "='01_Wejście'!B24"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Dochód po planowaniu rodziny [CHF/rok]'
    ws['B6'] = "='15_Planowanie_rodziny'!B52"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A7'] = 'Użyj dochodu skorygowanego? (0/1)'
    ws['B7'] = 0
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A8'] = 'Dochód użyty do analizy podatkowej [CHF/rok]'
    ws['B8'] = '=IF(B7=1,B6,B5)'
    set_cell_style(ws['A8'], font_bold=True)
    set_cell_style(ws['B8'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA B – Parametry nieruchomości i odliczeń
    # ========================================================================
    
    ws['A18'] = 'PARAMETRY NIERUCHOMOŚCI I ODLICZEŃ'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Wartość nieruchomości [CHF]'
    ws['B20'] = "='01_Wejście'!B4"
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A21'] = 'Odsetki miesięczne (rzeczywiste) [CHF]'
    ws['B21'] = "='04_Cashflow'!B2"
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A22'] = 'Odsetki roczne do odliczeń [CHF]'
    ws['B22'] = '=B21*12'
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A24'] = 'Eigenmietwert – % wartości nieruchomości'
    ws['B24'] = 0.025
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A25'] = 'Eigenmietwert roczny [CHF]'
    ws['B25'] = '=B20*B24'
    set_cell_style(ws['A25'])
    set_cell_style(ws['B25'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A26'] = 'Koszty utrzymania roczne do odliczenia [CHF]'
    ws['B26'] = ''
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='CCE5FF', number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA C – Tabela efektywnych stawek podatkowych
    # ========================================================================
    
    ws['A35'] = 'TABELA STAWEK EFEKTYWNYCH DLA WYBRANYCH KANTONÓW'
    set_cell_style(ws['A35'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli
    headers_tax = ['Kod kantonu', 'Gmina / miasto', 'Podatek dochodowy [%]', 
                   'Podatek majątkowy [%]', 'Uwagi']
    
    for col_idx, header in enumerate(headers_tax, start=1):
        cell = ws.cell(row=37, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Dane przykładowe (23 wiersze kantonów - użytkownik może edytować)
    canton_data = [
        ('ZH', 'Zürich', 0.180, 0.0020, 'Największe miasto'),
        ('ZG', 'Zug', 0.120, 0.0015, 'Niskie podatki'),
        ('BE', 'Bern', 0.210, 0.0025, 'Stolica'),
        ('GE', 'Genève', 0.240, 0.0030, 'Wysokie podatki'),
        ('VS', 'Sion', 0.150, 0.0018, 'Wallis'),
        ('TI', 'Lugano', 0.165, 0.0020, 'Ticino'),
        ('LU', 'Luzern', 0.175, 0.0022, 'Lucerna'),
        ('SZ', 'Schwyz', 0.130, 0.0012, 'Niskie podatki'),
        ('NW', 'Stans', 0.145, 0.0015, 'Nidwalden'),
        ('OW', 'Sarnen', 0.155, 0.0016, 'Obwalden'),
        ('UR', 'Altdorf', 0.160, 0.0017, 'Uri'),
        ('GL', 'Glarus', 0.165, 0.0019, 'Glarus'),
        ('ZG', 'Baar', 0.115, 0.0014, 'Zug - Baar'),
        ('FR', 'Fribourg', 0.195, 0.0023, 'Fryburg'),
        ('SO', 'Solothurn', 0.185, 0.0021, 'Solura'),
        ('BS', 'Basel', 0.220, 0.0028, 'Bazylea'),
        ('BL', 'Liestal', 0.200, 0.0024, 'Basel-Land'),
        ('SH', 'Schaffhausen', 0.175, 0.0020, 'Szafuza'),
        ('AR', 'Herisau', 0.170, 0.0019, 'Appenzell AR'),
        ('AI', 'Appenzell', 0.158, 0.0017, 'Appenzell AI'),
        ('SG', 'St. Gallen', 0.190, 0.0022, 'St. Gallen'),
        ('GR', 'Chur', 0.172, 0.0020, 'Gryzonia'),
        ('AG', 'Aarau', 0.188, 0.0023, 'Argowia'),
    ]
    
    for idx, (code, city, income_tax, wealth_tax, note) in enumerate(canton_data, start=38):
        ws.cell(row=idx, column=1).value = code
        set_cell_style(ws.cell(row=idx, column=1), bg_color='CCE5FF')
        
        ws.cell(row=idx, column=2).value = city
        set_cell_style(ws.cell(row=idx, column=2), bg_color='CCE5FF')
        
        ws.cell(row=idx, column=3).value = income_tax
        set_cell_style(ws.cell(row=idx, column=3), bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
        
        ws.cell(row=idx, column=4).value = wealth_tax
        set_cell_style(ws.cell(row=idx, column=4), bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
        
        ws.cell(row=idx, column=5).value = note
        set_cell_style(ws.cell(row=idx, column=5), bg_color='CCE5FF')
    
    # ========================================================================
    # SEKCJA D – Porównanie kantonów side-by-side
    # ========================================================================
    
    ws['A65'] = 'PORÓWNANIE KANTONÓW – DOCHÓD NETTO PO OPODATKOWANIU'
    set_cell_style(ws['A65'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki kolumn porównania
    comparison_headers = ['Parametr', 'Kanton 1', 'Kanton 2', 'Kanton 3', 'Kanton 4', 'Kanton 5']
    for col_idx, header in enumerate(comparison_headers, start=1):
        cell = ws.cell(row=67, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Wiersz wyboru kantonu (data validation)
    ws['A68'] = 'Kod kantonu'
    set_cell_style(ws['A68'], font_bold=True, bg_color='E0E0E0')
    
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Data validation dla wyboru kantonów
    dv = DataValidation(type="list", formula1='"ZH,ZG,BE,GE,VS,TI,LU,SZ,NW,OW,UR,GL,FR,SO,BS,BL,SH,AR,AI,SG,GR,AG"', 
                        allow_blank=True)
    ws.add_data_validation(dv)
    
    for col in range(2, 7):
        cell = ws.cell(row=68, column=col)
        cell.value = ''
        set_cell_style(cell, bg_color='CCE5FF')
        dv.add(cell)
    
    # Stawka podatku dochodowego
    ws['A69'] = 'Stawka podatku dochodowego [%]'
    set_cell_style(ws['A69'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}69'] = f'=IFERROR(INDEX($C$38:$C$60,MATCH({col_letter}68,$A$38:$A$60,0)),0)'
        set_cell_style(ws.cell(row=69, column=col), bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Stawka podatku majątkowego
    ws['A70'] = 'Stawka podatku majątkowego [%]'
    set_cell_style(ws['A70'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}70'] = f'=IFERROR(INDEX($D$38:$D$60,MATCH({col_letter}68,$A$38:$A$60,0)),0)'
        set_cell_style(ws.cell(row=70, column=col), bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Dochód bazowy
    ws['A72'] = 'Dochód bazowy [CHF/rok]'
    set_cell_style(ws['A72'])
    
    for col in range(2, 7):
        ws.cell(row=72, column=col).value = '=$B$8'
        set_cell_style(ws.cell(row=72, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    
    # Dochód opodatkowany
    ws['A73'] = 'Dochód opodatkowany = dochód + EMW – odsetki – koszty [CHF]'
    set_cell_style(ws['A73'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}73'] = f'=$B$8+$B$25-$B$22-$B$26'
        set_cell_style(ws.cell(row=73, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    
    # Podatek dochodowy
    ws['A75'] = 'Podatek dochodowy [CHF/rok]'
    set_cell_style(ws['A75'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}75'] = f'={col_letter}73*{col_letter}69'
        set_cell_style(ws.cell(row=75, column=col), bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    # Podatek majątkowy
    ws['A76'] = 'Podatek majątkowy [CHF/rok]'
    set_cell_style(ws['A76'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}76'] = f'=$B$20*{col_letter}70'
        set_cell_style(ws.cell(row=76, column=col), bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    # Suma podatków
    ws['A78'] = 'Suma podatków [CHF/rok]'
    set_cell_style(ws['A78'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}78'] = f'={col_letter}75+{col_letter}76'
        set_cell_style(ws.cell(row=78, column=col), bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Dochód netto
    ws['A79'] = 'Dochód netto po podatkach [CHF/rok]'
    set_cell_style(ws['A79'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}79'] = f'=$B$8-{col_letter}78'
        set_cell_style(ws.cell(row=79, column=col), bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Efektywna stopa podatku
    ws['A80'] = 'Efektywna stopa podatku [%]'
    set_cell_style(ws['A80'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}80'] = f'=IF($B$8>0,{col_letter}78/$B$8,0)'
        set_cell_style(ws.cell(row=80, column=col), bg_color='FFEB9C', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA E – Ranking kantonów
    # ========================================================================
    
    ws['A85'] = 'RANKING KANTONÓW'
    set_cell_style(ws['A85'], font_bold=True, font_size=12, border=False)
    
    ws['A87'] = 'Najwyższy dochód netto [CHF]'
    ws['B87'] = '=MAX(B79:F79)'
    set_cell_style(ws['A87'], font_bold=True)
    set_cell_style(ws['B87'], bg_color='C6EFCE', font_bold=True, number_format='#,##0.00')
    
    ws['A88'] = 'Najniższy dochód netto [CHF]'
    ws['B88'] = '=MIN(B79:F79)'
    set_cell_style(ws['A88'], font_bold=True)
    set_cell_style(ws['B88'], bg_color='FFC7CE', font_bold=True, number_format='#,##0.00')
    
    ws['A89'] = 'Różnica między najlepszym a najgorszym kantonem [CHF]'
    ws['B89'] = '=B87-B88'
    set_cell_style(ws['A89'], font_bold=True)
    set_cell_style(ws['B89'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A91'] = 'Najlepszy kanton (najwyższy dochód netto)'
    ws['B91'] = '=IFERROR(INDEX($B$68:$F$68,MATCH(B87,$B$79:$F$79,0)),"-")'
    set_cell_style(ws['A91'], font_bold=True)
    set_cell_style(ws['B91'], bg_color='C6EFCE', font_bold=True)
    
    ws['A92'] = 'Najgorszy kanton (najniższy dochód netto)'
    ws['B92'] = '=IFERROR(INDEX($B$68:$F$68,MATCH(B88,$B$79:$F$79,0)),"-")'
    set_cell_style(ws['A92'], font_bold=True)
    set_cell_style(ws['B92'], bg_color='FFC7CE', font_bold=True)
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

def create_renovation_sheet(wb):
    """Tworzy arkusz 16_Renowacje - model remontu i renowacji nieruchomości."""
    ws = wb.create_sheet('16_Renowacje')
    
    # Nagłówek główny
    ws['A1'] = 'MODEL REMONTU I RENOWACJI NIERUCHOMOŚCI'
    set_cell_style(ws['A1'], font_bold=True, font_size=14, border=False)
    
    # ========================================================================
    # SEKCJA A – Parametry ogólne
    # ========================================================================
    
    ws['A3'] = 'PARAMETRY OGÓLNE'
    set_cell_style(ws['A3'], font_bold=True, font_size=12, border=False)
    
    ws['A5'] = 'Wartość zakupu [CHF]'
    ws['B5'] = "='01_Wejście'!B4"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Horyzont analizy [lata]'
    ws['B6'] = 30
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format='0')
    
    ws['A7'] = 'Scenariusz wzrostu wartości (1=pess, 2=base, 3=opt)'
    ws['B7'] = 2
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A8'] = 'Procent kosztów inwestycyjnych (wpływających na wartość)'
    ws['B8'] = 0.7
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A9'] = 'Uwaga: Scenariusz 1=pesymistyczny, 2=bazowy, 3=optymistyczny (z arkusza 08)'
    set_cell_style(ws['A9'], border=False)
    set_cell_style(ws['B9'], border=False)
    
    # ========================================================================
    # SEKCJA B – Lista remontów
    # ========================================================================
    
    ws['A14'] = 'PLAN REMONTÓW / RENOWACJI'
    set_cell_style(ws['A14'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli remontów
    headers_reno = ['Nr', 'Opis remontu', 'Rok wykonania', 'Koszt [CHF]', 
                    'Czy zwiększa standard (0/1)', 'Procent inwestycyjny [%]']
    
    for col_idx, header in enumerate(headers_reno, start=1):
        cell = ws.cell(row=16, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Wiersze danych remontów (17-26 = 10 pozycji)
    for idx in range(10):
        row = 17 + idx
        
        # Nr
        ws.cell(row=row, column=1).value = idx + 1
        set_cell_style(ws.cell(row=row, column=1), bg_color='F2F2F2')
        
        # Opis
        ws.cell(row=row, column=2).value = ''
        set_cell_style(ws.cell(row=row, column=2), bg_color='CCE5FF')
        
        # Rok wykonania
        ws.cell(row=row, column=3).value = ''
        set_cell_style(ws.cell(row=row, column=3), bg_color='CCE5FF', number_format='0')
        
        # Koszt
        ws.cell(row=row, column=4).value = ''
        set_cell_style(ws.cell(row=row, column=4), bg_color='CCE5FF', number_format='#,##0.00')
        
        # Czy zwiększa standard
        ws.cell(row=row, column=5).value = ''
        set_cell_style(ws.cell(row=row, column=5), bg_color='CCE5FF', number_format='0')
        
        # Procent inwestycyjny
        ws.cell(row=row, column=6).value = f'=IF(E{row}=1,$B$8,0)'
        set_cell_style(ws.cell(row=row, column=6), bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA C – Roczna agregacja remontów
    # ========================================================================
    
    ws['A32'] = 'KOSZTY REMONTÓW W CZASIE'
    set_cell_style(ws['A32'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli agregacji
    headers_agg = ['Rok', 'Koszt roczny [CHF]', 'Część inwestycyjna [CHF]', 
                   'Część utrzymaniowa [CHF]', 'Skumulowany koszt [CHF]']
    
    for col_idx, header in enumerate(headers_agg, start=1):
        cell = ws.cell(row=34, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz bazowy dla skumulowanego)
    ws['A35'] = 0
    ws['B35'] = 0
    ws['C35'] = 0
    ws['D35'] = 0
    ws['E35'] = 0
    
    for col in range(1, 6):
        set_cell_style(ws.cell(row=35, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    ws.cell(row=35, column=1).number_format = '0'
    
    # Lata 1-30 (wiersze 36-65)
    for year in range(1, 31):
        row = 35 + year
        
        # Rok
        ws.cell(row=row, column=1).value = year
        ws.cell(row=row, column=1).number_format = '0'
        
        # Koszt roczny (SUMIF)
        ws.cell(row=row, column=2).value = f'=SUMIF($C$17:$C$26,A{row},$D$17:$D$26)'
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        # Część inwestycyjna (SUMPRODUCT)
        ws.cell(row=row, column=3).value = f'=SUMPRODUCT(($C$17:$C$26=A{row})*$D$17:$D$26*$F$17:$F$26)'
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # Część utrzymaniowa
        ws.cell(row=row, column=4).value = f'=B{row}-C{row}'
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        
        # Skumulowany koszt
        ws.cell(row=row, column=5).value = f'=E{row-1}+B{row}'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
    
    # ========================================================================
    # SEKCJA D – Wpływ remontów na wartość nieruchomości
    # ========================================================================
    
    ws['G32'] = 'WPŁYW REMONTÓW NA WARTOŚĆ NIERUCHOMOŚCI'
    set_cell_style(ws['G32'], font_bold=True, font_size=12, border=False)
    ws.merge_cells('G32:K32')
    
    # Nagłówki
    headers_value = ['Rok', 'Wartość wg scenariusza 08 [CHF]', 
                     'Skumulowana część inwestycyjna [CHF]', 
                     'Wartość po remontach [CHF]', 'Różnica vs scenariusz [CHF]']
    
    for col_idx, header in enumerate(headers_value, start=7):
        cell = ws.cell(row=34, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz 35)
    ws['G35'] = 0
    ws['H35'] = "=$B$5"
    ws['I35'] = 0
    ws['J35'] = '=H35+I35'
    ws['K35'] = '=J35-H35'
    
    for col in range(7, 12):
        set_cell_style(ws.cell(row=35, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    ws.cell(row=35, column=7).number_format = '0'
    
    # Lata 1-30 (wiersze 36-65)
    for year in range(1, 31):
        row = 35 + year
        sheet08_row = 13 + year  # rok 1 w arkuszu 08 to wiersz 14, itd.
        
        # Rok
        ws.cell(row=row, column=7).value = year
        ws.cell(row=row, column=7).number_format = '0'
        
        # Wartość wg scenariusza 08 (trójstopniowy IF)
        ws.cell(row=row, column=8).value = (
            f"=IF($B$7=1,'08_Symulacja_wzrostu_wartości'!B{sheet08_row},"
            f"IF($B$7=2,'08_Symulacja_wzrostu_wartości'!C{sheet08_row},"
            f"'08_Symulacja_wzrostu_wartości'!D{sheet08_row}))"
        )
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        
        # Skumulowana część inwestycyjna
        ws.cell(row=row, column=9).value = f'=I{row-1}+C{row}'
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        
        # Wartość po remontach
        ws.cell(row=row, column=10).value = f'=H{row}+I{row}'
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        
        # Różnica
        ws.cell(row=row, column=11).value = f'=J{row}-H{row}'
        ws.cell(row=row, column=11).number_format = '#,##0.00'
    
    # ========================================================================
    # SEKCJA E – Podsumowanie sprzedażowe
    # ========================================================================
    
    ws['A70'] = 'PODSUMOWANIE DLA MOMENTU SPRZEDAŻY'
    set_cell_style(ws['A70'], font_bold=True, font_size=12, border=False)
    
    ws['A72'] = 'Horyzont sprzedaży (X lat)'
    ws['B72'] = "='12_Analiza_sprzedaży_po_X_latach'!B7"
    set_cell_style(ws['A72'])
    set_cell_style(ws['B72'], bg_color='F2F2F2', font_bold=True, number_format='0')
    
    ws['A74'] = 'Skumulowany koszt remontów do roku X [CHF]'
    ws['B74'] = '=IF(B72<=30,INDEX($E$35:$E$65,B72+1),0)'
    set_cell_style(ws['A74'], font_bold=True)
    set_cell_style(ws['B74'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A75'] = 'Dodatkowa wartość dzięki remontom do roku X [CHF]'
    ws['B75'] = '=IF(B72<=30,INDEX($K$35:$K$65,B72+1),0)'
    set_cell_style(ws['A75'], font_bold=True)
    set_cell_style(ws['B75'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A76'] = 'Efekt netto remontów przy sprzedaży [CHF]'
    ws['B76'] = '=B75-B74'
    set_cell_style(ws['A76'], font_bold=True)
    set_cell_style(ws['B76'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A78'] = 'Interpretacja efektu netto'
    ws['B78'] = '=IF(B76>0,"Remonty zwiększają wartość netto","Remonty nie pokrywają kosztów")'
    set_cell_style(ws['A78'], font_bold=True)
    set_cell_style(ws['B78'], bg_color='FFEB9C', font_bold=True)
    
    # Formatowanie warunkowe dla B78
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    ws.conditional_formatting.add('B78',
        Rule(type='containsText', operator='containsText', text='zwiększają', 
             dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('B78',
        Rule(type='containsText', operator='containsText', text='nie pokrywają', 
             dxf=DifferentialStyle(fill=red_fill)))
    
    # ========================================================================
    # SEKCJA F – Statystyki dodatkowe
    # ========================================================================
    
    ws['A82'] = 'STATYSTYKI REMONTÓW'
    set_cell_style(ws['A82'], font_bold=True, font_size=12, border=False)
    
    ws['A84'] = 'Liczba planowanych remontów'
    ws['B84'] = '=COUNTIF($C$17:$C$26,">0")'
    set_cell_style(ws['A84'])
    set_cell_style(ws['B84'], bg_color='F2F2F2', number_format='0')
    
    ws['A85'] = 'Całkowity koszt wszystkich remontów [CHF]'
    ws['B85'] = '=SUM($D$17:$D$26)'
    set_cell_style(ws['A85'])
    set_cell_style(ws['B85'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A86'] = 'Średni koszt remontu [CHF]'
    ws['B86'] = '=IF(B84>0,B85/B84,0)'
    set_cell_style(ws['A86'])
    set_cell_style(ws['B86'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A87'] = 'Procent kosztów inwestycyjnych [%]'
    ws['B87'] = '=IF(B85>0,SUMPRODUCT($D$17:$D$26*$F$17:$F$26)/B85,0)'
    set_cell_style(ws['A87'])
    set_cell_style(ws['B87'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A88'] = 'Całkowity koszt inwestycyjny [CHF]'
    ws['B88'] = '=SUMPRODUCT($D$17:$D$26*$F$17:$F$26)'
    set_cell_style(ws['A88'])
    set_cell_style(ws['B88'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A89'] = 'Całkowity koszt utrzymaniowy [CHF]'
    ws['B89'] = '=B85-B88'
    set_cell_style(ws['A89'])
    set_cell_style(ws['B89'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18

def create_tax_canton_analysis_sheet(wb):
    """Tworzy arkusz 17_Podatki_kantony - analiza podatkowa kantonów."""
    ws = wb.create_sheet('17_Podatki_kantony')
    
    # Nagłówek główny
    ws['A1'] = 'ANALIZA PODATKOWA – PORÓWNANIE KANTONÓW'
    set_cell_style(ws['A1'], font_bold=True, font_size=14, border=False)
    
    # ========================================================================
    # SEKCJA A – Wybór dochodu (źródło)
    # ========================================================================
    
    ws['A3'] = 'DANE DOCHODOWE (ŹRÓDŁO)'
    set_cell_style(ws['A3'], font_bold=True, font_size=12, border=False)
    
    ws['A5'] = 'Dochód brutto roczny (bazowy) [CHF]'
    ws['B5'] = "='01_Wejście'!B24"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Dochód po planowaniu rodziny [CHF/rok]'
    ws['B6'] = "='15_Planowanie_rodziny'!B52"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A7'] = 'Użyj dochodu skorygowanego? (0/1)'
    ws['B7'] = 0
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A8'] = 'Dochód użyty do analizy podatkowej [CHF/rok]'
    ws['B8'] = '=IF(B7=1,B6,B5)'
    set_cell_style(ws['A8'], font_bold=True)
    set_cell_style(ws['B8'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA B – Parametry nieruchomości i odliczeń
    # ========================================================================
    
    ws['A18'] = 'PARAMETRY NIERUCHOMOŚCI I ODLICZEŃ'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Wartość nieruchomości [CHF]'
    ws['B20'] = "='01_Wejście'!B4"
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A21'] = 'Odsetki miesięczne (rzeczywiste) [CHF]'
    ws['B21'] = "='04_Cashflow'!B2"
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A22'] = 'Odsetki roczne do odliczeń [CHF]'
    ws['B22'] = '=B21*12'
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A24'] = 'Eigenmietwert – % wartości nieruchomości'
    ws['B24'] = 0.025
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A25'] = 'Eigenmietwert roczny [CHF]'
    ws['B25'] = '=B20*B24'
    set_cell_style(ws['A25'])
    set_cell_style(ws['B25'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A26'] = 'Koszty utrzymania roczne do odliczenia [CHF]'
    ws['B26'] = ''
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='CCE5FF', number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA C – Tabela efektywnych stawek podatkowych
    # ========================================================================
    
    ws['A35'] = 'TABELA STAWEK EFEKTYWNYCH DLA WYBRANYCH KANTONÓW'
    set_cell_style(ws['A35'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli
    headers_tax = ['Kod kantonu', 'Gmina / miasto', 'Podatek dochodowy [%]', 
                   'Podatek majątkowy [%]', 'Uwagi']
    
    for col_idx, header in enumerate(headers_tax, start=1):
        cell = ws.cell(row=37, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Dane przykładowe (23 wiersze kantonów - użytkownik może edytować)
    canton_data = [
        ('ZH', 'Zürich', 0.180, 0.0020, 'Największe miasto'),
        ('ZG', 'Zug', 0.120, 0.0015, 'Niskie podatki'),
        ('BE', 'Bern', 0.210, 0.0025, 'Stolica'),
        ('GE', 'Genève', 0.240, 0.0030, 'Wysokie podatki'),
        ('VS', 'Sion', 0.150, 0.0018, 'Wallis'),
        ('TI', 'Lugano', 0.165, 0.0020, 'Ticino'),
        ('LU', 'Luzern', 0.175, 0.0022, 'Lucerna'),
        ('SZ', 'Schwyz', 0.130, 0.0012, 'Niskie podatki'),
        ('NW', 'Stans', 0.145, 0.0015, 'Nidwalden'),
        ('OW', 'Sarnen', 0.155, 0.0016, 'Obwalden'),
        ('UR', 'Altdorf', 0.160, 0.0017, 'Uri'),
        ('GL', 'Glarus', 0.165, 0.0019, 'Glarus'),
        ('ZG', 'Baar', 0.115, 0.0014, 'Zug - Baar'),
        ('FR', 'Fribourg', 0.195, 0.0023, 'Fryburg'),
        ('SO', 'Solothurn', 0.185, 0.0021, 'Solura'),
        ('BS', 'Basel', 0.220, 0.0028, 'Bazylea'),
        ('BL', 'Liestal', 0.200, 0.0024, 'Basel-Land'),
        ('SH', 'Schaffhausen', 0.175, 0.0020, 'Szafuza'),
        ('AR', 'Herisau', 0.170, 0.0019, 'Appenzell AR'),
        ('AI', 'Appenzell', 0.158, 0.0017, 'Appenzell AI'),
        ('SG', 'St. Gallen', 0.190, 0.0022, 'St. Gallen'),
        ('GR', 'Chur', 0.172, 0.0020, 'Gryzonia'),
        ('AG', 'Aarau', 0.188, 0.0023, 'Argowia'),
    ]
    
    for idx, (code, city, income_tax, wealth_tax, note) in enumerate(canton_data, start=38):
        ws.cell(row=idx, column=1).value = code
        set_cell_style(ws.cell(row=idx, column=1), bg_color='CCE5FF')
        
        ws.cell(row=idx, column=2).value = city
        set_cell_style(ws.cell(row=idx, column=2), bg_color='CCE5FF')
        
        ws.cell(row=idx, column=3).value = income_tax
        set_cell_style(ws.cell(row=idx, column=3), bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
        
        ws.cell(row=idx, column=4).value = wealth_tax
        set_cell_style(ws.cell(row=idx, column=4), bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
        
        ws.cell(row=idx, column=5).value = note
        set_cell_style(ws.cell(row=idx, column=5), bg_color='CCE5FF')
    
    # ========================================================================
    # SEKCJA D – Porównanie kantonów side-by-side
    # ========================================================================
    
    ws['A65'] = 'PORÓWNANIE KANTONÓW – DOCHÓD NETTO PO OPODATKOWANIU'
    set_cell_style(ws['A65'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki kolumn porównania
    comparison_headers = ['Parametr', 'Kanton 1', 'Kanton 2', 'Kanton 3', 'Kanton 4', 'Kanton 5']
    for col_idx, header in enumerate(comparison_headers, start=1):
        cell = ws.cell(row=67, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Wiersz wyboru kantonu (data validation)
    ws['A68'] = 'Kod kantonu'
    set_cell_style(ws['A68'], font_bold=True, bg_color='E0E0E0')
    
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Data validation dla wyboru kantonów
    dv = DataValidation(type="list", formula1='"ZH,ZG,BE,GE,VS,TI,LU,SZ,NW,OW,UR,GL,FR,SO,BS,BL,SH,AR,AI,SG,GR,AG"', 
                        allow_blank=True)
    ws.add_data_validation(dv)
    
    for col in range(2, 7):
        cell = ws.cell(row=68, column=col)
        cell.value = ''
        set_cell_style(cell, bg_color='CCE5FF')
        dv.add(cell)
    
    # Stawka podatku dochodowego
    ws['A69'] = 'Stawka podatku dochodowego [%]'
    set_cell_style(ws['A69'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}69'] = f'=IFERROR(INDEX($C$38:$C$60,MATCH({col_letter}68,$A$38:$A$60,0)),0)'
        set_cell_style(ws.cell(row=69, column=col), bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Stawka podatku majątkowego
    ws['A70'] = 'Stawka podatku majątkowego [%]'
    set_cell_style(ws['A70'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}70'] = f'=IFERROR(INDEX($D$38:$D$60,MATCH({col_letter}68,$A$38:$A$60,0)),0)'
        set_cell_style(ws.cell(row=70, column=col), bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Dochód bazowy
    ws['A72'] = 'Dochód bazowy [CHF/rok]'
    set_cell_style(ws['A72'])
    
    for col in range(2, 7):
        ws.cell(row=72, column=col).value = '=$B$8'
        set_cell_style(ws.cell(row=72, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    
    # Dochód opodatkowany
    ws['A73'] = 'Dochód opodatkowany = dochód + EMW – odsetki – koszty [CHF]'
    set_cell_style(ws['A73'])
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}73'] = f'=$B$8+$B$25-$B$22-$B$26'
        set_cell_style(ws.cell(row=73, column=col), bg_color='F2F2F2', number_format='#,##0.00')
    
    # Podatek dochodowy
    ws['A75'] = 'Podatek dochodowy [CHF/rok]'
    set_cell_style(ws['A75'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}75'] = f'={col_letter}73*{col_letter}69'
        set_cell_style(ws.cell(row=75, column=col), bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    # Podatek majątkowy
    ws['A76'] = 'Podatek majątkowy [CHF/rok]'
    set_cell_style(ws['A76'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}76'] = f'=$B$20*{col_letter}70'
        set_cell_style(ws.cell(row=76, column=col), bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    # Suma podatków
    ws['A78'] = 'Suma podatków [CHF/rok]'
    set_cell_style(ws['A78'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}78'] = f'={col_letter}75+{col_letter}76'
        set_cell_style(ws.cell(row=78, column=col), bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Dochód netto
    ws['A79'] = 'Dochód netto po podatkach [CHF/rok]'
    set_cell_style(ws['A79'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}79'] = f'=$B$8-{col_letter}78'
        set_cell_style(ws.cell(row=79, column=col), bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Efektywna stopa podatku
    ws['A80'] = 'Efektywna stopa podatku [%]'
    set_cell_style(ws['A80'], font_bold=True)
    
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}80'] = f'=IF($B$8>0,{col_letter}78/$B$8,0)'
        set_cell_style(ws.cell(row=80, column=col), bg_color='FFEB9C', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    # ========================================================================
    # SEKCJA E – Ranking kantonów
    # ========================================================================
    
    ws['A85'] = 'RANKING KANTONÓW'
    set_cell_style(ws['A85'], font_bold=True, font_size=12, border=False)
    
    ws['A87'] = 'Najwyższy dochód netto [CHF]'
    ws['B87'] = '=MAX(B79:F79)'
    set_cell_style(ws['A87'], font_bold=True)
    set_cell_style(ws['B87'], bg_color='C6EFCE', font_bold=True, number_format='#,##0.00')
    
    ws['A88'] = 'Najniższy dochód netto [CHF]'
    ws['B88'] = '=MIN(B79:F79)'
    set_cell_style(ws['A88'], font_bold=True)
    set_cell_style(ws['B88'], bg_color='FFC7CE', font_bold=True, number_format='#,##0.00')
    
    ws['A89'] = 'Różnica między najlepszym a najgorszym kantonem [CHF]'
    ws['B89'] = '=B87-B88'
    set_cell_style(ws['A89'], font_bold=True)
    set_cell_style(ws['B89'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A91'] = 'Najlepszy kanton (najwyższy dochód netto)'
    ws['B91'] = '=IFERROR(INDEX($B$68:$F$68,MATCH(B87,$B$79:$F$79,0)),"-")'
    set_cell_style(ws['A91'], font_bold=True)
    set_cell_style(ws['B91'], bg_color='C6EFCE', font_bold=True)
    
    ws['A92'] = 'Najgorszy kanton (najniższy dochód netto)'
    ws['B92'] = '=IFERROR(INDEX($B$68:$F$68,MATCH(B88,$B$79:$F$79,0)),"-")'
    set_cell_style(ws['A92'], font_bold=True)
    set_cell_style(ws['B92'], bg_color='FFC7CE', font_bold=True)
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


def create_liquidity_and_buffer_sheet(wb):
    """Tworzy arkusz 18_Plynnosc_poduszka - analiza płynności i poduszki finansowej."""
    ws = wb.create_sheet('18_Plynnosc_poduszka')
    
    # Nagłówek główny
    ws['A1'] = 'ANALIZA PŁYNNOŚCI I PODUSZKI FINANSOWEJ'
    set_cell_style(ws['A1'], font_bold=True, font_size=14, border=False)
    
    # ========================================================================
    # SEKCJA A – Dane bazowe (dochód, koszty, poduszka)
    # ========================================================================
    
    ws['A3'] = 'DANE BAZOWE'
    set_cell_style(ws['A3'], font_bold=True, font_size=12, border=False)
    
    # Dochód
    ws['A5'] = 'Dochód roczny użyty w modelu [CHF]'
    ws['B5'] = "='15_Planowanie_rodziny'!B57"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Dochód miesięczny [CHF]'
    ws['B6'] = '=B5/12'
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    # Koszty
    ws['A8'] = 'Łączny miesięczny koszt posiadania mieszkania [CHF]'
    ws['B8'] = "='04_Cashflow'!B8"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'Inne koszty życia miesięcznie (jedzenie, transport, ubezpieczenia) [CHF]'
    ws['B9'] = ''
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A10'] = 'Inne kredyty / leasing / alimenty [CHF/mies]'
    ws['B10'] = ''
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='CCE5FF', number_format='#,##0.00')
    
    # Poduszka i płynne aktywa
    ws['A12'] = 'Gotówka + płynne inwestycje (poduszka finansowa) [CHF]'
    ws['B12'] = ''
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A13'] = 'Dostępne limity kredytowe / overdraft jako awaryjny bufor [CHF]'
    ws['B13'] = ''
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A14'] = 'Łączny bufor płynności [CHF]'
    ws['B14'] = '=B12+B13'
    set_cell_style(ws['A14'], font_bold=True)
    set_cell_style(ws['B14'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # ========================================================================
    # SEKCJA B – Bazowa płynność (bez szoku)
    # ========================================================================
    
    ws['A17'] = 'BAZOWA PŁYNNOŚĆ (BEZ SZOKU)'
    set_cell_style(ws['A17'], font_bold=True, font_size=12, border=False)
    
    ws['A19'] = 'Łączne wydatki stałe miesięcznie [CHF]'
    ws['B19'] = '=B8+B9+B10'
    set_cell_style(ws['A19'], font_bold=True)
    set_cell_style(ws['B19'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A20'] = 'Nadwyżka (lub deficyt) miesięczna [CHF]'
    ws['B20'] = '=B6-B19'
    set_cell_style(ws['A20'], font_bold=True)
    set_cell_style(ws['B20'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    # Formatowanie warunkowe dla nadwyżki/deficytu
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    ws.conditional_formatting.add('B20', CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    ws.conditional_formatting.add('B20', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    ws['A22'] = 'Ile miesięcy przetrwasz na poduszce przy zerowym dochodzie (pełne wydatki stałe)?'
    ws['B22'] = '=IF(B19>0,B14/B19,0)'
    set_cell_style(ws['A22'], font_bold=True)
    set_cell_style(ws['B22'], bg_color='FFEB9C', font_bold=True, number_format='0.0')
    
    # ========================================================================
    # SEKCJA C – Runway przy częściowym spadku dochodu
    # ========================================================================
    
    ws['A25'] = 'PŁYNNOŚĆ PRZY CZĘŚCIOWYM SPADKU DOCHODU'
    set_cell_style(ws['A25'], font_bold=True, font_size=12, border=False)
    
    ws['A27'] = 'Scenariusz: dochód spada do X% obecnego'
    ws['B27'] = 0.5
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A28'] = 'Deficyt miesięczny przy dochodzie X% [CHF]'
    ws['B28'] = '=MAX(0,B19-B6*B27)'
    set_cell_style(ws['A28'])
    set_cell_style(ws['B28'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A29'] = 'Ile miesięcy starczy poduszki przy dochodzie X%'
    ws['B29'] = '=IF(B28>0,B14/B28,0)'
    set_cell_style(ws['A29'], font_bold=True)
    set_cell_style(ws['B29'], bg_color='F2F2F2', font_bold=True, number_format='0.0')
    
    # ========================================================================
    # SEKCJA D – Cele poduszki (3 / 6 / 12 miesięcy)
    # ========================================================================
    
    ws['A32'] = 'CEL PODUSZKI (3 / 6 / 12 MIESIĘCY)'
    set_cell_style(ws['A32'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli
    headers_goals = ['Scenariusz', 'Miesiące', 'Wymagana poduszka [CHF]', 
                     'Nadwyżka / niedobór [CHF]', 'Status']
    
    for col_idx, header in enumerate(headers_goals, start=1):
        cell = ws.cell(row=34, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Wiersz 1: Minimalna (3 mies.)
    ws['A35'] = 'Minimalna (3 mies.)'
    ws['B35'] = 3
    ws['C35'] = '=B35*$B$19'
    ws['D35'] = '=$B$14-C35'
    ws['E35'] = '=IF(D35>=0,"OK","Brakuje")'
    
    set_cell_style(ws['A35'])
    set_cell_style(ws['B35'], bg_color='F2F2F2', number_format='0')
    set_cell_style(ws['C35'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['D35'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['E35'], bg_color='F2F2F2', font_bold=True)
    
    # Wiersz 2: Konserwatywna (6 mies.)
    ws['A36'] = 'Konserwatywna (6 mies.)'
    ws['B36'] = 6
    ws['C36'] = '=B36*$B$19'
    ws['D36'] = '=$B$14-C36'
    ws['E36'] = '=IF(D36>=0,"OK","Brakuje")'
    
    set_cell_style(ws['A36'])
    set_cell_style(ws['B36'], bg_color='F2F2F2', number_format='0')
    set_cell_style(ws['C36'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['D36'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['E36'], bg_color='F2F2F2', font_bold=True)
    
    # Wiersz 3: Bardzo konserwatywna (12 mies.)
    ws['A37'] = 'Bardzo konserwatywna (12 mies.)'
    ws['B37'] = 12
    ws['C37'] = '=B37*$B$19'
    ws['D37'] = '=$B$14-C37'
    ws['E37'] = '=IF(D37>=0,"OK","Brakuje")'
    
    set_cell_style(ws['A37'])
    set_cell_style(ws['B37'], bg_color='F2F2F2', number_format='0')
    set_cell_style(ws['C37'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['D37'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['E37'], bg_color='F2F2F2', font_bold=True)
    
    # Formatowanie warunkowe dla kolumny D i E
    ws.conditional_formatting.add('D35:D37', CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=green_fill))
    ws.conditional_formatting.add('D35:D37', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    ws.conditional_formatting.add('E35:E37',
        Rule(type='containsText', operator='containsText', text='OK', 
             dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('E35:E37',
        Rule(type='containsText', operator='containsText', text='Brakuje', 
             dxf=DifferentialStyle(fill=red_fill)))
    
    # ========================================================================
    # SEKCJA E – Scenariusze kryzysowe
    # ========================================================================
    
    ws['A41'] = 'SCENARIUSZE KRYZYSOWE – TEST PODUSZKI'
    set_cell_style(ws['A41'], font_bold=True, font_size=12, border=False)
    
    # Nagłówki tabeli
    headers_crisis = ['Scenariusz', 'Spadek dochodu [%]', 'Liczba miesięcy', 
                      'Dodatkowe koszty miesięczne [CHF]', 'Deficyt miesięczny [CHF]',
                      'Łączny deficyt [CHF]', 'Czy poduszka wystarczy?', 'Margines po scenariuszu [CHF]']
    
    for col_idx, header in enumerate(headers_crisis, start=1):
        cell = ws.cell(row=43, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Scenariusz 1: Spadek 50% na 6 miesięcy
    ws['A44'] = 'Spadek dochodu o 50% na 6 miesięcy'
    ws['B44'] = 0.5
    ws['C44'] = 6
    ws['D44'] = 0
    ws['E44'] = '=MAX(0,$B$19-($B$6*(1-B44))+D44)'
    ws['F44'] = '=E44*C44'
    ws['G44'] = '=IF($B$14>=F44,"TAK","NIE")'
    ws['H44'] = '=$B$14-F44'
    
    set_cell_style(ws['A44'])
    set_cell_style(ws['B44'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    set_cell_style(ws['C44'], bg_color='CCE5FF', number_format='0')
    set_cell_style(ws['D44'], bg_color='CCE5FF', number_format='#,##0.00')
    set_cell_style(ws['E44'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['F44'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['G44'], bg_color='F2F2F2', font_bold=True)
    set_cell_style(ws['H44'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Scenariusz 2: Utrata 100% dochodu na 3 miesiące
    ws['A45'] = 'Spadek dochodu o 100% na 3 mies. + dodatkowe koszty'
    ws['B45'] = 1.0
    ws['C45'] = 3
    ws['D45'] = 0
    ws['E45'] = '=MAX(0,$B$19-($B$6*(1-B45))+D45)'
    ws['F45'] = '=E45*C45'
    ws['G45'] = '=IF($B$14>=F45,"TAK","NIE")'
    ws['H45'] = '=$B$14-F45'
    
    set_cell_style(ws['A45'])
    set_cell_style(ws['B45'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    set_cell_style(ws['C45'], bg_color='CCE5FF', number_format='0')
    set_cell_style(ws['D45'], bg_color='CCE5FF', number_format='#,##0.00')
    set_cell_style(ws['E45'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['F45'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['G45'], bg_color='F2F2F2', font_bold=True)
    set_cell_style(ws['H45'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Scenariusz 3: Duży jednorazowy wydatek
    ws['A46'] = 'Duży jednorazowy wydatek [CHF]'
    ws['B46'] = ''
    ws['C46'] = ''
    ws['D46'] = ''
    ws['E46'] = ''
    ws['F46'] = '=IF(D46>0,D46,0)'
    ws['G46'] = '=IF(D46>0,IF($B$14>=F46,"TAK","NIE"),"")'
    ws['H46'] = '=IF(D46>0,$B$14-F46,"")'
    
    set_cell_style(ws['A46'])
    set_cell_style(ws['B46'], bg_color='F2F2F2')
    set_cell_style(ws['C46'], bg_color='F2F2F2')
    set_cell_style(ws['D46'], bg_color='CCE5FF', number_format='#,##0.00')
    set_cell_style(ws['E46'], bg_color='F2F2F2')
    set_cell_style(ws['F46'], bg_color='F2F2F2', number_format='#,##0.00')
    set_cell_style(ws['G46'], bg_color='F2F2F2', font_bold=True)
    set_cell_style(ws['H46'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Formatowanie warunkowe dla kolumny G i H
    ws.conditional_formatting.add('G44:G46',
        Rule(type='containsText', operator='containsText', text='TAK', 
             dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('G44:G46',
        Rule(type='containsText', operator='containsText', text='NIE', 
             dxf=DifferentialStyle(fill=red_fill)))
    
    ws.conditional_formatting.add('H44:H46', CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=green_fill))
    ws.conditional_formatting.add('H44:H46', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    # ========================================================================
    # SEKCJA F – Wskaźniki podsumowujące
    # ========================================================================
    
    ws['A51'] = 'PODSUMOWANIE PŁYNNOŚCI'
    set_cell_style(ws['A51'], font_bold=True, font_size=12, border=False)
    
    ws['A53'] = 'Poduszka / roczny dochód [%]'
    ws['B53'] = '=IF(B5>0,B14/(B5*12),0)'  # ← POPRAWIONA FORMUŁA
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A54'] = 'Poduszka / miesięczne wydatki stałe [miesiące]'
    ws['B54'] = '=IF(B19>0,B14/B19,0)'
    set_cell_style(ws['A54'], font_bold=True)
    set_cell_style(ws['B54'], bg_color='F2F2F2', font_bold=True, number_format='0.0')
    
    ws['A56'] = 'Werdykt ogólny'
    ws['B56'] = ('=IF(B22>=6,"Bezpieczna poduszka (≥6 mies. pełnych wydatków)",'
                 'IF(B22>=3,"OK, ale warto ją powiększyć",'
                 '"Zbyt mała poduszka – priorytet: oszczędzanie"))')
    set_cell_style(ws['A56'], font_bold=True)
    set_cell_style(ws['B56'], bg_color='FFF2CC', font_bold=True)
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18


def main():
    """Główna funkcja tworząca cały skoroszyt z 18 arkuszami."""
    print("Tworzenie rozszerzonego kalkulatora nieruchomości w Szwajcarii...")
    
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    print("  -> Tworzenie arkusza 00_Stałe...")
    create_constants_sheet(wb)
    
    print("  -> Tworzenie arkusza 01_Wejście...")
    create_input_sheet(wb)
    
    print("  -> Tworzenie arkusza 02_Finansowanie...")
    create_financing_sheet(wb)
    
    print("  -> Tworzenie arkusza 03_Tragbarkeit...")
    create_tragbarkeit_sheet(wb)
    
    print("  -> Tworzenie arkusza 04_Cashflow...")
    create_cashflow_sheet(wb)
    
    print("  -> Tworzenie arkusza 05_Harmonogram_roczny...")
    create_yearly_schedule_sheet(wb)
    
    print("  -> Tworzenie arkusza 06_Harmonogram_miesieczny...")
    create_monthly_schedule_sheet(wb)
    
    print("  -> Tworzenie arkusza 07_Analiza_ROI...")
    create_roi_sheet(wb)
    
    print("  -> Tworzenie arkusza 08_Symulacja_wzrostu_wartości...")
    create_appreciation_sheet(wb)
    
    print("  -> Tworzenie arkusza 09_Koszt_alternatywny_kapitalu...")
    create_opportunity_cost_sheet(wb)
    
    print("  -> Tworzenie arkusza 10_Rent_vs_Buy_30lat...")
    create_rent_vs_buy_sheet(wb)
    
    print("  -> Tworzenie arkusza 11_Stress_test...")
    create_stress_test_sheet(wb)
    
    print("  -> Tworzenie arkusza 12_Analiza_sprzedaży_po_X_latach...")
    create_sale_analysis_sheet(wb)
    
    print("  -> Tworzenie arkusza 13_Analiza_PRD...")
    create_prd_analysis_sheet(wb)
    
    print("  -> Tworzenie arkusza 14_Nowa_nieruchomosc_X_lat...")
    create_new_property_after_sale_sheet(wb)
    
    print("  -> Tworzenie arkusza 15_Planowanie_rodziny...")
    create_family_planning_sheet(wb)
    
    print("  -> Tworzenie arkusza 16_Renowacje...")
    create_renovation_sheet(wb)
    
    print("  -> Tworzenie arkusza 17_Podatki_kantony...")
    create_tax_canton_analysis_sheet(wb)
    
    print("  -> Tworzenie arkusza 18_Plynnosc_poduszka...")
    create_liquidity_and_buffer_sheet(wb)
    
    wb.active = wb['01_Wejście']
    
    filename = 'kalkulator_nieruchomosc_CH.xlsx'
    wb.save(filename)
    
    print(f"\n✅ Plik '{filename}' został utworzony pomyślnie!")
    print("\nStruktura arkuszy:")
    print("  00_Stałe - Parametry ogólne")
    print("  01_Wejście - Dane wejściowe użytkownika")
    print("  02_Finansowanie - Analiza kredytu")
    print("  03_Tragbarkeit - Test zdolności kredytowej")
    print("  04_Cashflow - Rzeczywiste koszty miesięczne")
    print("  05_Harmonogram_roczny - Harmonogram spłat rocznych (30 lat)")
    print("  06_Harmonogram_miesieczny - Harmonogram spłat miesięcznych (360 miesięcy)")
    print("  07_Analiza_ROI - Budowa equity i ROI w czasie")
    print("  08_Symulacja_wzrostu_wartości - 3 scenariusze wzrostu wartości nieruchomości")
    print("  09_Koszt_alternatywny_kapitalu - Porównanie equity z alternatywną inwestycją (ETF)")
    print("  10_Rent_vs_Buy_30lat - Porównanie kupna vs wynajmu w horyzoncie 30 lat")
    print("  11_Stress_test - Szok stóp procentowych a koszty i Tragbarkeit")
    print("  12_Analiza_sprzedaży_po_X_latach - Analiza wyniku sprzedaży po X latach, z uwzględnieniem spłaty kredytu")
    print("  13_Analiza_PRD - Price-to-Rent Ratio, yield i interpretacja wyceny")
    print("  14_Nowa_nieruchomosc_X_lat - Analiza maksymalnej ceny nowej nieruchomości po sprzedaży obecnej")
    print("  15_Planowanie_rodziny - Symulacja zmian dochodu po narodzinach dzieci")
    print("  16_Renowacje - Model remontu i renowacji nieruchomości")
    print("  17_Podatki_kantony - Analiza podatkowa kantonów")
    print("  18_Plynnosc_poduszka - Analiza płynności i poduszki finansowej")
    print("\nInstrukcja użytkowania:")
    print("1. Otwórz plik w LibreOffice Calc")
    print("2. Przejdź do arkusza '01_Wejście'")
    print("3. Wypełnij niebieskie pola danymi (w tym dobrowolną amortyzację H1)")
    print("4. W arkuszu 07_Analiza_ROI wpisz szacowany roczny wzrost wartości")
    print("5. W arkuszu 08_Symulacja wpisz 3 scenariusze wzrostu")
    print("6. W arkuszu 09_Koszt_alternatywny wpisz stopę zwrotu ETF")
    print("7. W arkuszu 10_Rent_vs_Buy wpisz wzrost czynszu i kosztów")
    print("8. Arkusz 11_Stress_test pokazuje wpływ wzrostu stóp procentowych")
    print("9. W arkuszu 12_Analiza_sprzedaży wpisz horyzont sprzedaży i wzrost wartości")
    print("10. Arkusz 13_Analiza_PRD pokazuje wskaźniki wyceny i yield")
    print("11. W arkuszu 14_Nowa_nieruchomosc_X_lat wpisz parametry przyszłego zakupu (wzrost dochodu, oszczędności, cenę testową)")
    print("12. W arkuszu 15_Planowanie_rodziny wpisz dane o dzieciach, etatach i kosztach opieki")
    print("13. W arkuszu 16_Renowacje zaplanuj remonty (rok, koszt, typ)")
    print("14. W arkuszu 17_Podatki_kantony wybierz 5 kantonów i porównaj obciążenia podatkowe")
    print("15. W arkuszu 18_Plynnosc_poduszka wpisz poduszkę finansową i inne koszty życia")
    print("16. Wyniki pojawią się automatycznie we wszystkich arkuszach")
    print("\nPowodzenia w analizie opłacalności zakupu nieruchomości! 🏠🇨🇭")


if __name__ == '__main__':
    main()