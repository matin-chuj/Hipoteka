# -*- coding: utf-8 -*-
"""
Skrypt tworzący rozszerzony kalkulator opłacalności zakupu nieruchomości w Szwajcarii.
Wersja z harmonogramami rocznymi i miesięcznymi oraz dobrowolną amortyzacją H1.

INSTRUKCJA UŻYCIA (Linux Mint):
1. Zainstaluj openpyxl: 
   sudo apt install python3-openpyxl
   
2. Uruchom skrypt:
   python3 build_kalkulator_nieruchomosc_ch.py
   
3. Plik kalkulator_nieruchomosc_CH.xlsx zostanie utworzony w bieżącym katalogu

4. Otwórz plik w LibreOffice Calc:
   libreoffice --calc kalkulator_nieruchomosc_CH.xlsx

AUTOR: GitHub Copilot
DATA: 2025-11-20
UŻYTKOWNIK: matin-chuj
SYSTEM: Linux Mint
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00


def set_cell_style(cell, font_bold=False, font_size=11, bg_color=None, 
                   border=True, number_format=None, alignment='left'):
    """Pomocnicza funkcja do ustawiania stylu komórki."""
    if font_bold or font_size != 11:
        cell.font = Font(name='Calibri', size=font_size, bold=font_bold)
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    if border:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    if number_format:
        cell.number_format = number_format
    
    if alignment == 'center':
        cell.alignment = Alignment(horizontal='center', vertical='center')
    elif alignment == 'right':
        cell.alignment = Alignment(horizontal='right', vertical='center')
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center')


def create_constants_sheet(wb):
    """Tworzy arkusz 00_Stałe z parametrami ogólnymi."""
    ws = wb.create_sheet('00_Stałe', 0)
    
    ws['A1'] = 'Parametr'
    ws['B1'] = 'Wartość domyślna'
    ws['C1'] = 'Opis'
    
    for cell in ['A1', 'B1', 'C1']:
        set_cell_style(ws[cell], font_bold=True)
    
    data = [
        ('Min. wkład własny ogółem', 0.20, 'Minimum 20% wartości nieruchomości'),
        ('Min. wkład własny gotówkowy', 0.10, 'Minimum 10% z gotówki (nie z filarów)'),
        ('LTV docelowe po amortyzacji', 0.65, 'Loan-to-Value po spłacie Hypoteki 2'),
        ('Lata amortyzacji do 65%', 15, 'Liczba lat na amortyzację'),
        ('Oprocentowanie testowe (bank)', 0.05, 'Stopa używana przez bank do testu zdolności'),
        ('Roczne koszty utrzymania (test)', 0.01, '1% wartości nieruchomości rocznie'),
        ('Max. Tragbarkeit (udział dochodu)', 0.33, 'Maksymalny udział kosztów w dochodzie'),
        ('Kurs CHF/PLN', 4.60, 'Aktualny kurs franka szwajcarskiego'),
        ('Procent kosztów transakcyjnych (notariusz itd.)', 0.025, 'Szacunkowe koszty notarialne i opłaty'),
    ]
    
    for idx, (param, value, desc) in enumerate(data, start=2):
        ws[f'A{idx}'] = param
        ws[f'B{idx}'] = value
        ws[f'C{idx}'] = desc
        set_cell_style(ws[f'A{idx}'])
        set_cell_style(ws[f'C{idx}'])
        
        if idx in [2, 3, 4, 6, 7, 8, 10]:
            set_cell_style(ws[f'B{idx}'], number_format=FORMAT_PERCENTAGE_00)
        elif idx == 5:
            set_cell_style(ws[f'B{idx}'], number_format='0')
        elif idx == 9:
            set_cell_style(ws[f'B{idx}'], number_format=FORMAT_NUMBER_00)
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50


def create_input_sheet(wb):
    """Tworzy arkusz 01_Wejście z danymi wejściowymi użytkownika."""
    ws = wb.create_sheet('01_Wejście')
    
    ws['A2'] = 'INFORMACJE O NIERUCHOMOŚCI'
    set_cell_style(ws['A2'], font_bold=True, font_size=12, border=False)
    
    ws['A4'] = 'Cena zakupu nieruchomości [CHF]'
    ws['B4'] = ''
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A5'] = 'Szacunkowy % kosztów transakcyjnych'
    ws['B5'] = "='00_Stałe'!B10"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'TWÓJ WKŁAD WŁASNY'
    set_cell_style(ws['A7'], font_bold=True, font_size=12, border=False)
    
    ws['A8'] = 'Wkład gotówkowy [CHF]'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A9'] = 'Wkład z II filaru [CHF]'
    ws['B9'] = ''
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A10'] = 'Wkład z III filaru [CHF]'
    ws['B10'] = ''
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A12'] = 'Suma wkładu własnego [CHF]'
    ws['B12'] = '=B8+B9+B10'
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A13'] = 'Min. wymagany wkład ogółem [CHF]'
    ws['B13'] = "=B4*'00_Stałe'!B2"
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Min. wymagany wkład gotówkowy [CHF]'
    ws['B14'] = "=B4*'00_Stałe'!B3"
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A15'] = 'Status wkładu ogółem'
    ws['B15'] = '=IF(B12>=B13,"OK","ZA MAŁO WKŁADU")'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', font_bold=True)
    
    ws['A16'] = 'Status wkładu gotówkowego'
    ws['B16'] = '=IF(B8>=B14,"OK","ZA MAŁO GOTÓWKI")'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', font_bold=True)
    
    ws['A18'] = 'DANE KREDYTOWE'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A19'] = 'Stopa % Hypoteka 1 (do 65%)'
    ws['B19'] = ''
    set_cell_style(ws['A19'])
    set_cell_style(ws['B19'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A20'] = 'Stopa % Hypoteka 2 (powyżej 65%)'
    ws['B20'] = ''
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A21'] = 'Rodzaj amortyzacji (D – bezpośr.; N – pośrednia)'
    ws['B21'] = ''
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='CCE5FF')
    
    dv = DataValidation(type="list", formula1='"D,N"', allow_blank=True)
    dv.error = 'Wprowadź D lub N'
    dv.errorTitle = 'Nieprawidłowa wartość'
    ws.add_data_validation(dv)
    dv.add(ws['B21'])
    
    ws['A23'] = 'TWOJE FINANSE'
    set_cell_style(ws['A23'], font_bold=True, font_size=12, border=False)
    
    ws['A24'] = 'Dochód brutto gospodarstwa roczny [CHF]'
    ws['B24'] = ''
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A25'] = 'Dobrowolna amortyzacja Hypoteki 1 – rocznie [CHF]'
    ws['B25'] = ''
    set_cell_style(ws['A25'])
    set_cell_style(ws['B25'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A26'] = 'Roczne koszty wspólnoty / Nebenkosten [CHF]'
    ws['B26'] = ''
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws['A27'] = 'Miesięczny czynsz przy wynajmie porównywalnego lokalu'
    ws['B27'] = ''
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='CCE5FF', number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.freeze_panes = 'A3'


def create_financing_sheet(wb):
    """Tworzy arkusz 02_Finansowanie z analizą kredytu."""
    ws = wb.create_sheet('02_Finansowanie')
    
    ws['A1'] = 'PODSTAWY FINANSOWANIA'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Cena zakupu [CHF]'
    ws['B2'] = "='01_Wejście'!B4"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Suma wkładu własnego [CHF]'
    ws['B3'] = "='01_Wejście'!B12"
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A4'] = 'Kwota kredytu (łącznie) [CHF]'
    ws['B4'] = '=B2-B3'
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'LTV początkowe'
    ws['B5'] = '=B4/B2'
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Hypoteka 1 (do 65% wartości) [CHF]'
    ws['B7'] = "=MIN(B4,B2*'00_Stałe'!B4)"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A8'] = 'Hypoteka 2 (powyżej 65%) [CHF]'
    ws['B8'] = '=B4-B7'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A10'] = 'Stopa % Hypoteka 1 (aktualna)'
    ws['B10'] = "='01_Wejście'!B19"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A11'] = 'Stopa % Hypoteka 2 (aktualna)'
    ws['B11'] = "='01_Wejście'!B20"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A13'] = 'Odsetki roczne H1 [CHF]'
    ws['B13'] = '=B7*B10'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Odsetki roczne H2 [CHF]'
    ws['B14'] = '=B8*B11'
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A15'] = 'Razem odsetki roczne [CHF]'
    ws['B15'] = '=B13+B14'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A16'] = 'Odsetki miesięczne [CHF]'
    ws['B16'] = '=B15/12'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A18'] = 'Kwota do amortyzacji (Hypoteka 2) [CHF]'
    ws['B18'] = '=B8'
    set_cell_style(ws['A18'])
    set_cell_style(ws['B18'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A19'] = 'Lata amortyzacji (do 65% LTV)'
    ws['B19'] = "='00_Stałe'!B5"
    set_cell_style(ws['A19'])
    set_cell_style(ws['B19'], bg_color='F2F2F2', font_bold=True, number_format='0')
    
    ws['A20'] = 'Amortyzacja roczna [CHF]'
    ws['B20'] = '=B18/B19'
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A21'] = 'Amortyzacja miesięczna [CHF]'
    ws['B21'] = '=B20/12'
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A22'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B22'] = "='01_Wejście'!B25"
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A23'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B23'] = '=B22/12'
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20


def create_tragbarkeit_sheet(wb):
    """Tworzy arkusz 03_Tragbarkeit z analizą zdolności kredytowej."""
    ws = wb.create_sheet('03_Tragbarkeit')
    
    ws['A1'] = 'TEST ZDOLNOŚCI KREDYTOWEJ (TRAGBARKEIT)'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Kwota kredytu łącznie [CHF]'
    ws['B2'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Stopa testowa banku'
    ws['B3'] = "='00_Stałe'!B6"
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A4'] = 'Odsetki testowe roczne [CHF]'
    ws['B4'] = '=B2*B3'
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'Cena zakupu [CHF]'
    ws['B5'] = "='01_Wejście'!B4"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Koszty utrzymania roczne (test) [CHF]'
    ws['B6'] = "=B5*'00_Stałe'!B7"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A7'] = 'Amortyzacja roczna (test) [CHF]'
    ws['B7'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A9'] = 'Łączne koszty roczne wg banku [CHF]'
    ws['B9'] = '=B4+B6+B7'
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A10'] = 'Dochód brutto roczny [CHF]'
    ws['B10'] = "='01_Wejście'!B24"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A11'] = 'Tragbarkeit (udział dochodu)'
    ws['B11'] = '=B9/B10'
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A12'] = 'Maksymalny udział dochodu'
    ws['B12'] = "='00_Stałe'!B8"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_PERCENTAGE_00)
    
    ws['A13'] = 'Ocena Tragbarkeit'
    ws['B13'] = '=IF(B11<=B12,"OK","ZA WYSOKIE OBCIĄŻENIE")'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True)
    
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add('B13', CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill))
    ws.conditional_formatting.add('B13', CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20


def create_cashflow_sheet(wb):
    """Tworzy arkusz 04_Cashflow z analizą rzeczywistych kosztów miesięcznych."""
    ws = wb.create_sheet('04_Cashflow')
    
    ws['A1'] = 'RZECZYWISTE KOSZTY MIESIĘCZNE'
    set_cell_style(ws['A1'], font_bold=True, font_size=12, border=False)
    
    ws['A2'] = 'Odsetki miesięczne (aktualne) [CHF]'
    ws['B2'] = "='02_Finansowanie'!B16"
    set_cell_style(ws['A2'])
    set_cell_style(ws['B2'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A3'] = 'Amortyzacja miesięczna – cash-out [CHF]'
    ws['B3'] = '=IF(\'01_Wejście\'!B21="D",\'02_Finansowanie\'!B21,0)+\'02_Finansowanie\'!B23'
    set_cell_style(ws['A3'])
    set_cell_style(ws['B3'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A4'] = 'Roczne koszty utrzymania realne [CHF]'
    ws['B4'] = "='01_Wejście'!B4*'00_Stałe'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczne koszty utrzymania [CHF]'
    ws['B5'] = '=B4/12'
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A6'] = 'Miesięczne koszty wspólnoty (HOA/NK) [CHF]'
    ws['B6'] = "='01_Wejście'!B26/12"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A8'] = 'Łączny miesięczny koszt posiadania [CHF]'
    ws['B8'] = '=B2+B3+B5+B6'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A9'] = 'Kurs CHF/PLN'
    ws['B9'] = "='00_Stałe'!B9"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', font_bold=True, number_format=FORMAT_NUMBER_00)
    
    ws['A10'] = 'Łączny miesięczny koszt posiadania [PLN]'
    ws['B10'] = '=B8*B9'
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='FFEB9C', font_bold=True, number_format='#,##0.00')
    
    ws['A12'] = 'Miesięczny czynsz przy wynajmie [CHF]'
    ws['B12'] = "='01_Wejście'!B27"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A13'] = 'Różnica: wynajem – posiadanie [CHF/mies.]'
    ws['B13'] = '=B12-B8'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', font_bold=True, number_format='#,##0.00')
    
    ws['A14'] = 'Komentarz'
    ws['B14'] = '=IF(B13>0,"Kupno tańsze od wynajmu","Kupno droższe od wynajmu")'
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2', font_bold=True)
    
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add('B14', Rule(type='containsText', operator='containsText', text='tańsze', dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('B14', Rule(type='containsText', operator='containsText', text='droższe', dxf=DifferentialStyle(fill=red_fill)))
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20


def create_yearly_schedule_sheet(wb):
    """Tworzy arkusz 05_Harmonogram_roczny z harmonogramem spłat rocznych."""
    ws = wb.create_sheet('05_Harmonogram_roczny')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'H1 początkowe [CHF]'
    ws['B4'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'H2 początkowe [CHF]'
    ws['B5'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Stopa H1 roczna'
    ws['B6'] = "='01_Wejście'!B19"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Stopa H2 roczna'
    ws['B7'] = "='01_Wejście'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Amortyzacja H2 roczna (obowiązkowa) [CHF]'
    ws['B8'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B9'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A10'] = 'Typ amortyzacji H2 (D/N)'
    ws['B10'] = "='01_Wejście'!B21"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2')
    
    headers = ['Rok', 'Saldo pocz. H1', 'Saldo pocz. H2', 'Saldo pocz. razem', 'Odsetki H1', 'Odsetki H2', 'Odsetki razem',
               'Amortyzacja H2', 'Amortyzacja H1', 'Amortyzacja razem', 'Cash-out roczny', 'Saldo końc. H1', 'Saldo końc. H2', 'Saldo końc. razem']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A13'] = 0
    ws['B13'] = '=$B$4'
    ws['C13'] = '=$B$5'
    ws['D13'] = '=B13+C13'
    ws['E13'] = 0
    ws['F13'] = 0
    ws['G13'] = 0
    ws['H13'] = 0
    ws['I13'] = 0
    ws['J13'] = 0
    ws['K13'] = 0
    ws['L13'] = '=B13'
    ws['M13'] = '=C13'
    ws['N13'] = '=L13+M13'
    
    for col in range(1, 15):
        cell = ws.cell(row=13, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    ws['A14'] = '=A13+1'
    ws['B14'] = '=L13'
    ws['C14'] = '=M13'
    ws['D14'] = '=B14+C14'
    ws['E14'] = '=B14*$B$6'
    ws['F14'] = '=C14*$B$7'
    ws['G14'] = '=E14+F14'
    ws['H14'] = '=IF($B$10="D",MIN($B$8,C14),0)'
    ws['I14'] = '=MIN($B$9,B14)'
    ws['J14'] = '=H14+I14'
    ws['K14'] = '=G14+J14'
    ws['L14'] = '=MAX(0,B14-I14)'
    ws['M14'] = '=MAX(0,C14-H14)'
    ws['N14'] = '=L14+M14'
    
    for col in range(1, 15):
        cell = ws.cell(row=14, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    for row in range(15, 45):
        for col in range(1, 15):
            if col == 1:
                ws.cell(row=row, column=col).value = f'=A{row-1}+1'
            elif col == 2:
                ws.cell(row=row, column=col).value = f'=L{row-1}'
            elif col == 3:
                ws.cell(row=row, column=col).value = f'=M{row-1}'
            elif col == 4:
                ws.cell(row=row, column=col).value = f'=B{row}+C{row}'
            elif col == 5:
                ws.cell(row=row, column=col).value = f'=B{row}*$B$6'
            elif col == 6:
                ws.cell(row=row, column=col).value = f'=C{row}*$B$7'
            elif col == 7:
                ws.cell(row=row, column=col).value = f'=E{row}+F{row}'
            elif col == 8:
                ws.cell(row=row, column=col).value = f'=IF($B$10="D",MIN($B$8,C{row}),0)'
            elif col == 9:
                ws.cell(row=row, column=col).value = f'=MIN($B$9,B{row})'
            elif col == 10:
                ws.cell(row=row, column=col).value = f'=H{row}+I{row}'
            elif col == 11:
                ws.cell(row=row, column=col).value = f'=G{row}+J{row}'
            elif col == 12:
                ws.cell(row=row, column=col).value = f'=MAX(0,B{row}-I{row})'
            elif col == 13:
                ws.cell(row=row, column=col).value = f'=MAX(0,C{row}-H{row})'
            elif col == 14:
                ws.cell(row=row, column=col).value = f'=L{row}+M{row}'
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 14


def create_monthly_schedule_sheet(wb):
    """Tworzy arkusz 06_Harmonogram_miesieczny z harmonogramem spłat miesięcznych."""
    ws = wb.create_sheet('06_Harmonogram_miesieczny')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'H1 początkowe [CHF]'
    ws['B4'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'H2 początkowe [CHF]'
    ws['B5'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Stopa H1 roczna'
    ws['B6'] = "='01_Wejście'!B19"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Stopa H2 roczna'
    ws['B7'] = "='01_Wejście'!B20"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Stopa H1 miesięczna'
    ws['B8'] = '=B6/12'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A9'] = 'Stopa H2 miesięczna'
    ws['B9'] = '=B7/12'
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A10'] = 'Amortyzacja H2 roczna [CHF]'
    ws['B10'] = "='02_Finansowanie'!B20"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A11'] = 'Amortyzacja H2 miesięczna [CHF]'
    ws['B11'] = "='02_Finansowanie'!B21"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A12'] = 'Dobrowolna amortyzacja H1 roczna [CHF]'
    ws['B12'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A13'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B13'] = "='02_Finansowanie'!B23"
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A14'] = 'Typ amortyzacji H2 (D/N)'
    ws['B14'] = "='01_Wejście'!B21"
    set_cell_style(ws['A14'])
    set_cell_style(ws['B14'], bg_color='F2F2F2')
    
    headers = ['Miesiąc', 'Saldo pocz. H1', 'Saldo pocz. H2', 'Saldo pocz. razem', 'Odsetki H1', 'Odsetki H2', 'Odsetki razem',
               'Amortyzacja H2', 'Amortyzacja H1', 'Amortyzacja razem', 'Cash-out miesięczny', 'Saldo końc. H1', 'Saldo końc. H2', 'Saldo końc. razem']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=18, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A19'] = 0
    ws['B19'] = '=$B$4'
    ws['C19'] = '=$B$5'
    ws['D19'] = '=B19+C19'
    ws['E19'] = 0
    ws['F19'] = 0
    ws['G19'] = 0
    ws['H19'] = 0
    ws['I19'] = 0
    ws['J19'] = 0
    ws['K19'] = 0
    ws['L19'] = '=B19'
    ws['M19'] = '=C19'
    ws['N19'] = '=L19+M19'
    
    for col in range(1, 15):
        cell = ws.cell(row=19, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    ws['A20'] = '=A19+1'
    ws['B20'] = '=L19'
    ws['C20'] = '=M19'
    ws['D20'] = '=B20+C20'
    ws['E20'] = '=B20*$B$8'
    ws['F20'] = '=C20*$B$9'
    ws['G20'] = '=E20+F20'
    ws['H20'] = '=IF($B$14="D",MIN($B$11,C20),0)'
    ws['I20'] = '=MIN($B$13,B20)'
    ws['J20'] = '=H20+I20'
    ws['K20'] = '=G20+J20'
    ws['L20'] = '=MAX(0,B20-I20)'
    ws['M20'] = '=MAX(0,C20-H20)'
    ws['N20'] = '=L20+M20'
    
    for col in range(1, 15):
        cell = ws.cell(row=20, column=col)
        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            cell.number_format = '#,##0.00'
    
    for row in range(21, 381):
        for col in range(1, 15):
            if col == 1:
                ws.cell(row=row, column=col).value = f'=A{row-1}+1'
            elif col == 2:
                ws.cell(row=row, column=col).value = f'=L{row-1}'
            elif col == 3:
                ws.cell(row=row, column=col).value = f'=M{row-1}'
            elif col == 4:
                ws.cell(row=row, column=col).value = f'=B{row}+C{row}'
            elif col == 5:
                ws.cell(row=row, column=col).value = f'=B{row}*$B$8'
            elif col == 6:
                ws.cell(row=row, column=col).value = f'=C{row}*$B$9'
            elif col == 7:
                ws.cell(row=row, column=col).value = f'=E{row}+F{row}'
            elif col == 8:
                ws.cell(row=row, column=col).value = f'=IF($B$14="D",MIN($B$11,C{row}),0)'
            elif col == 9:
                ws.cell(row=row, column=col).value = f'=MIN($B$13,B{row})'
            elif col == 10:
                ws.cell(row=row, column=col).value = f'=H{row}+I{row}'
            elif col == 11:
                ws.cell(row=row, column=col).value = f'=G{row}+J{row}'
            elif col == 12:
                ws.cell(row=row, column=col).value = f'=MAX(0,B{row}-I{row})'
            elif col == 13:
                ws.cell(row=row, column=col).value = f'=MAX(0,C{row}-H{row})'
            elif col == 14:
                ws.cell(row=row, column=col).value = f'=L{row}+M{row}'
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 14


def create_roi_sheet(wb):
    """Tworzy arkusz 07_Analiza_ROI z analizą budowy equity i ROI."""
    ws = wb.create_sheet('07_Analiza_ROI')
    
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wartość zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Wkład własny początkowy [CHF]'
    ws['B5'] = "='01_Wejście'!B12"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Kwota kredytu łącznie [CHF]'
    ws['B6'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Saldo początkowe H1 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Saldo początkowe H2 [CHF]'
    ws['B8'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'Kurs CHF/PLN'
    ws['B9'] = "='00_Stałe'!B9"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format=FORMAT_NUMBER_00)
    
    ws['A10'] = 'Roczny wzrost wartości nieruchomości'
    ws['B10'] = ''
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    headers = ['Rok', 'Wartość nieruchomości [CHF]', 'Saldo H1 [CHF]', 'Saldo H2 [CHF]',
               'Saldo razem [CHF]', 'Equity [CHF]', 'Przyrost equity r/r [CHF]',
               'Skumulowany przyrost equity [CHF]', 'ROI z wkładu (%)']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    ws['A16'] = 0
    ws['B16'] = '=$B$4'
    ws['C16'] = '=$B$7'
    ws['D16'] = '=$B$8'
    ws['E16'] = '=C16+D16'
    ws['F16'] = '=B16-E16'
    ws['G16'] = 0
    ws['H16'] = 0
    ws['I16'] = '=F16/$B$5'
    
    for col in range(2, 9):
        cell = ws.cell(row=16, column=col)
        if col <= 8:
            cell.number_format = '#,##0.00'
    ws.cell(row=16, column=9).number_format = FORMAT_PERCENTAGE_00
    
    ws['A17'] = '=A16+1'
    ws['B17'] = '=B16*(1+$B$10)'
    ws['C17'] = "='05_Harmonogram_roczny'!L14"
    ws['D17'] = "='05_Harmonogram_roczny'!M14"
    ws['E17'] = '=C17+D17'
    ws['F17'] = '=B17-E17'
    ws['G17'] = '=F17-F16'
    ws['H17'] = '=H16+G17'
    ws['I17'] = '=F17/$B$5'
    
    for col in range(2, 9):
        cell = ws.cell(row=17, column=col)
        if col <= 8:
            cell.number_format = '#,##0.00'
    ws.cell(row=17, column=9).number_format = FORMAT_PERCENTAGE_00
    
    for row in range(18, 47):
        year = row - 16
        harmonogram_row = 13 + year
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$10)'
        ws.cell(row=row, column=3).value = f"='05_Harmonogram_roczny'!L{harmonogram_row}"
        ws.cell(row=row, column=4).value = f"='05_Harmonogram_roczny'!M{harmonogram_row}"
        ws.cell(row=row, column=5).value = f'=C{row}+D{row}'
        ws.cell(row=row, column=6).value = f'=B{row}-E{row}'
        ws.cell(row=row, column=7).value = f'=F{row}-F{row-1}'
        ws.cell(row=row, column=8).value = f'=H{row-1}+G{row}'
        ws.cell(row=row, column=9).value = f'=F{row}/$B$5'
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            if col <= 8:
                cell.number_format = '#,##0.00'
        ws.cell(row=row, column=9).number_format = FORMAT_PERCENTAGE_00
    
    ws['A50'] = 'Equity po 30 latach [CHF]'
    ws['B50'] = '=F46'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'ROI całkowite z wkładu'
    ws['B51'] = '=I46'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A52'] = 'Średni roczny zwrot CAGR'
    ws['B52'] = '=(1+I46)^(1/30)-1'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 16

# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD create_roi_sheet (przed main)
# ============================================================================

def create_appreciation_sheet(wb):
    """Tworzy arkusz 08_Symulacja_wzrostu_wartości z 3 scenariuszami."""
    ws = wb.create_sheet('08_Symulacja_wzrostu_wartości')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wartość początkowa nieruchomości [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Scenariusz pesymistyczny – roczny wzrost'
    ws['B6'] = ''
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Scenariusz bazowy – roczny wzrost'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Scenariusz optymistyczny – roczny wzrost'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    # Nagłówki tabeli
    headers = ['Rok', 'Pesymistyczny [CHF]', 'Bazowy [CHF]', 'Optymistyczny [CHF]']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0
    ws['A13'] = 0
    ws['B13'] = '=$B$4'
    ws['C13'] = '=$B$4'
    ws['D13'] = '=$B$4'
    
    for col in range(2, 5):
        ws.cell(row=13, column=col).number_format = '#,##0.00'
    
    # Rok 1 (szablon)
    ws['A14'] = '=A13+1'
    ws['B14'] = '=B13*(1+$B$6)'
    ws['C14'] = '=C13*(1+$B$7)'
    ws['D14'] = '=D13*(1+$B$8)'
    
    for col in range(2, 5):
        ws.cell(row=14, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 43)
    for row in range(15, 44):
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$6)'
        ws.cell(row=row, column=3).value = f'=C{row-1}*(1+$B$7)'
        ws.cell(row=row, column=4).value = f'=D{row-1}*(1+$B$8)'
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A46'] = 'Wartość po 30 latach – pesymistyczny'
    ws['B46'] = '=B43'
    set_cell_style(ws['A46'], font_bold=True)
    set_cell_style(ws['B46'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A47'] = 'Wartość po 30 latach – bazowy'
    ws['B47'] = '=C43'
    set_cell_style(ws['A47'], font_bold=True)
    set_cell_style(ws['B47'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A48'] = 'Wartość po 30 latach – optymistyczny'
    ws['B48'] = '=D43'
    set_cell_style(ws['A48'], font_bold=True)
    set_cell_style(ws['B48'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A50'] = 'CAGR scenariusz bazowy'
    ws['B50'] = '=(C43/$B$4)^(1/30)-1'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_appreciation_sheet (przed main)
# ============================================================================

def create_opportunity_cost_sheet(wb):
    """Tworzy arkusz 09_Koszt_alternatywny_kapitalu - porównanie ETF vs equity."""
    ws = wb.create_sheet('09_Koszt_alternatywny_kapitalu')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Wkład własny początkowy [CHF]'
    ws['B4'] = "='01_Wejście'!B12"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Roczna stopa zwrotu alternatywnej inwestycji'
    ws['B5'] = ''
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A6'] = 'Roczna kwota odpowiadająca amortyzacji H2 [CHF]'
    ws['B6'] = '=IF(\'01_Wejście\'!B21="D",\'02_Finansowanie\'!B20,0)'
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Roczna kwota odpowiadająca dobrowolnej amortyzacji H1 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B22"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Łączna roczna kwota dodatkowych wpłat [CHF]'
    ws['B8'] = '=B6+B7'
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli
    headers = ['Rok', 'Kapitał pocz. w roku [CHF]', 'Nowe wpłaty w roku [CHF]', 
               'Zysk z inwestycji w roku [CHF]', 'Kapitał końcowy [CHF]', 
               'Equity w nieruchomości [CHF]', 'Różnica: ETF – equity [CHF]', 'Komentarz']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz 16)
    ws['A16'] = 0
    ws['B16'] = '=$B$4'
    ws['C16'] = 0
    ws['D16'] = 0
    ws['E16'] = '=B16+C16+D16'
    ws['F16'] = "='07_Analiza_ROI'!F16"
    ws['G16'] = '=E16-F16'
    ws['H16'] = '=IF(G16>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
    
    for col in range(2, 8):
        ws.cell(row=16, column=col).number_format = '#,##0.00'
    
    # Rok 1 (wiersz 17 - szablon)
    ws['A17'] = '=A16+1'
    ws['B17'] = '=E16'
    ws['C17'] = '=$B$8'
    ws['D17'] = '=(B17+C17)*$B$5'
    ws['E17'] = '=B17+C17+D17'
    ws['F17'] = "='07_Analiza_ROI'!F17"
    ws['G17'] = '=E17-F17'
    ws['H17'] = '=IF(G17>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
    
    for col in range(2, 8):
        ws.cell(row=17, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 46)
    for row in range(18, 47):
        year = row - 16
        roi_row = 16 + year
        
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=E{row-1}'
        ws.cell(row=row, column=3).value = '=$B$8'
        ws.cell(row=row, column=4).value = f'=(B{row}+C{row})*$B$5'
        ws.cell(row=row, column=5).value = f'=B{row}+C{row}+D{row}'
        ws.cell(row=row, column=6).value = f"='07_Analiza_ROI'!F{roi_row}"
        ws.cell(row=row, column=7).value = f'=E{row}-F{row}'
        ws.cell(row=row, column=8).value = f'=IF(G{row}>0,"ETF > nieruchomość","ETF ≤ nieruchomość")'
        
        for col in range(2, 8):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A50'] = 'Kapitał alternatywny po 30 latach [CHF]'
    ws['B50'] = '=E46'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'Equity w nieruchomości po 30 latach [CHF]'
    ws['B51'] = '=F46'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A52'] = 'Różnica: ETF – nieruchomość [CHF]'
    ws['B52'] = '=B50-B51'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A53'] = 'Komentarz'
    ws['B53'] = '=IF(B52>0,"Lepsza inwestycja w ETF","Lepsza inwestycja w nieruchomość")'
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['H'].width = 25



# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD definicją create_opportunity_cost_sheet (lub ostatnią funkcją tworzącą arkusze)
# ============================================================================

def create_rent_vs_buy_sheet(wb):
    """Tworzy arkusz 10_Rent_vs_Buy_30lat - porównanie kupna vs wynajmu w 30 latach."""
    ws = wb.create_sheet('10_Rent_vs_Buy_30lat')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Miesięczny koszt posiadania [CHF]'
    ws['B4'] = "='04_Cashflow'!B8"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczny czynsz wynajmu [CHF]'
    ws['B5'] = "='01_Wejście'!B27"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Roczny wzrost czynszu'
    ws['B6'] = ''
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A7'] = 'Roczny wzrost kosztów posiadania'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A8'] = 'Equity po 30 latach [CHF]'
    ws['B8'] = "='07_Analiza_ROI'!F46"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli porównania
    headers = ['Rok', 'Koszt posiadania (Kupno) – roczny [CHF]', 'Koszt wynajmu – roczny [CHF]',
               'Różnica (Wynajem – Kupno) [CHF]', 'Skumulowana różnica [CHF]']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Rok 0 (wiersz 16)
    ws['A16'] = 0
    ws['B16'] = '=$B$4*12'
    ws['C16'] = '=$B$5*12'
    ws['D16'] = '=C16-B16'
    ws['E16'] = '=D16'
    
    for col in range(2, 6):
        ws.cell(row=16, column=col).number_format = '#,##0.00'
    
    # Rok 1 (wiersz 17 - szablon)
    ws['A17'] = '=A16+1'
    ws['B17'] = '=B16*(1+$B$7)'
    ws['C17'] = '=C16*(1+$B$6)'
    ws['D17'] = '=C17-B17'
    ws['E17'] = '=E16+D17'
    
    for col in range(2, 6):
        ws.cell(row=17, column=col).number_format = '#,##0.00'
    
    # Kopiowanie do roku 30 (wiersz 46)
    for row in range(18, 47):
        ws.cell(row=row, column=1).value = f'=A{row-1}+1'
        ws.cell(row=row, column=2).value = f'=B{row-1}*(1+$B$7)'
        ws.cell(row=row, column=3).value = f'=C{row-1}*(1+$B$6)'
        ws.cell(row=row, column=4).value = f'=C{row}-B{row}'
        ws.cell(row=row, column=5).value = f'=E{row-1}+D{row}'
        
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Podsumowanie
    ws['A50'] = 'Suma kosztów posiadania przez 30 lat [CHF]'
    ws['B50'] = '=SUM(B16:B46)'
    set_cell_style(ws['A50'], font_bold=True)
    set_cell_style(ws['B50'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A51'] = 'Suma kosztów wynajmu przez 30 lat [CHF]'
    ws['B51'] = '=SUM(C16:C46)'
    set_cell_style(ws['A51'], font_bold=True)
    set_cell_style(ws['B51'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A52'] = 'Różnica (wynajem – kupno) [CHF]'
    ws['B52'] = '=B51-B50'
    set_cell_style(ws['A52'], font_bold=True)
    set_cell_style(ws['B52'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A53'] = 'Equity po 30 latach [CHF]'
    ws['B53'] = '=$B$8'
    set_cell_style(ws['A53'], font_bold=True)
    set_cell_style(ws['B53'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A54'] = 'Efekt netto (różnica + equity) [CHF]'
    ws['B54'] = '=B52+B53'
    set_cell_style(ws['A54'], font_bold=True)
    set_cell_style(ws['B54'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A55'] = 'Komentarz'
    ws['B55'] = '=IF(B54>0,"Kupno opłaca się bardziej","Wynajem opłaca się bardziej")'
    set_cell_style(ws['A55'], font_bold=True)
    set_cell_style(ws['B55'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_rent_vs_buy_sheet
# ============================================================================

def create_stress_test_sheet(wb):
    """Tworzy arkusz 11_Stress_test - szok stóp procentowych a koszty i Tragbarkeit."""
    ws = wb.create_sheet('11_Stress_test')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Stopa H1 aktualna'
    ws['B4'] = "='01_Wejście'!B19"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A5'] = 'Stopa H2 aktualna'
    ws['B5'] = "='01_Wejście'!B20"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A6'] = 'Saldo H1 [CHF]'
    ws['B6'] = "='02_Finansowanie'!B7"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Saldo H2 [CHF]'
    ws['B7'] = "='02_Finansowanie'!B8"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A8'] = 'Koszt utrzymania miesięczny [CHF]'
    ws['B8'] = "='04_Cashflow'!B5"
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A9'] = 'HOA/Nebenkosten miesięczne [CHF]'
    ws['B9'] = "='04_Cashflow'!B6"
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A10'] = 'Amortyzacja H2 miesięczna [CHF]'
    ws['B10'] = "='02_Finansowanie'!B21"
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A11'] = 'Dobrowolna amortyzacja H1 miesięczna [CHF]'
    ws['B11'] = "='02_Finansowanie'!B23"
    set_cell_style(ws['A11'])
    set_cell_style(ws['B11'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A12'] = 'Miesięczny czynsz wynajmu [CHF]'
    ws['B12'] = "='01_Wejście'!B27"
    set_cell_style(ws['A12'])
    set_cell_style(ws['B12'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Nagłówki tabeli scenariuszy
    headers = ['Szok stopy (Δ, w p.p.)', 'Odsetki H1 miesięczne [CHF]', 'Odsetki H2 miesięczne [CHF]',
               'Łączne odsetki miesięczne [CHF]', 'Miesięczny cash-out (posiadanie) [CHF]',
               'Różnica vs wynajem [CHF]', 'Tragbarkeit (udział dochodu)', 'Status']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color='D0D0D0', alignment='center')
    
    # Scenariusze szoku
    shocks = [0, 0.005, 0.01, 0.015, 0.02]
    
    for idx, shock in enumerate(shocks):
        row = 16 + idx
        
        # Szok w kolumnie A
        ws.cell(row=row, column=1).value = shock
        ws.cell(row=row, column=1).number_format = FORMAT_PERCENTAGE_00
        
        # B: Odsetki H1 miesięczne
        ws.cell(row=row, column=2).value = f'=$A$6*(($A$4+A{row})/12)'
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        # C: Odsetki H2 miesięczne
        ws.cell(row=row, column=3).value = f'=$A$7*(($A$5+A{row})/12)'
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # D: Łączne odsetki
        ws.cell(row=row, column=4).value = f'=B{row}+C{row}'
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        
        # E: Miesięczny cash-out
        ws.cell(row=row, column=5).value = f'=D{row}+$A$8+$A$9+$A$10+$A$11'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        
        # F: Różnica vs wynajem
        ws.cell(row=row, column=6).value = f'=$A$12-E{row}'
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        # G: Tragbarkeit
        ws.cell(row=row, column=7).value = f"=(E{row}*12)/'01_Wejście'!B24"
        ws.cell(row=row, column=7).number_format = FORMAT_PERCENTAGE_00
        
        # H: Status
        ws.cell(row=row, column=8).value = f"=IF(G{row}<'00_Stałe'!B8,\"OK\",\"Ryzyko Tragbarkeit\")"
    
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 22
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20


# ============================================================================
# 1) NOWA FUNKCJA - wkleić POD create_stress_test_sheet (lub ostatnią funkcją tworzącą arkusze)
# ============================================================================

def create_sale_analysis_sheet(wb):
    """Tworzy arkusz 12_Analiza_sprzedaży_po_X_latach - analiza wyniku sprzedaży po X latach."""
    ws = wb.create_sheet('12_Analiza_sprzedaży_po_X_latach')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Cena zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Wkład własny początkowy [CHF]'
    ws['B5'] = "='01_Wejście'!B12"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Kwota kredytu łącznie [CHF]'
    ws['B6'] = "='02_Finansowanie'!B4"
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Lata do sprzedaży (X)'
    ws['B7'] = ''
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='CCE5FF', number_format='0')
    
    ws['A8'] = 'Roczny wzrost wartości nieruchomości'
    ws['B8'] = ''
    set_cell_style(ws['A8'])
    set_cell_style(ws['B8'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A9'] = 'Prowizja/koszty sprzedaży [% od ceny]'
    ws['B9'] = ''
    set_cell_style(ws['A9'])
    set_cell_style(ws['B9'], bg_color='CCE5FF', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A10'] = 'Rok sprzedaży w harmonogramie'
    ws['B10'] = '=13+$B$7'
    set_cell_style(ws['A10'])
    set_cell_style(ws['B10'], bg_color='F2F2F2', number_format='0')
    
    # Wycena nieruchomości przy sprzedaży
    ws['A13'] = 'WYCENA NIERUCHOMOŚCI PRZY SPRZEDAŻY'
    set_cell_style(ws['A13'], font_bold=True, font_size=12, border=False)
    
    ws['A15'] = 'Wartość nieruchomości przy sprzedaży [CHF]'
    ws['B15'] = '=B4*(1+$B$8)^$B$7'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A16'] = 'Koszty sprzedaży (prowizje, podatki) [CHF]'
    ws['B16'] = '=B15*$B$9'
    set_cell_style(ws['A16'])
    set_cell_style(ws['B16'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Saldo kredytu przy sprzedaży
    ws['A18'] = 'SALDO KREDYTU PRZY SPRZEDAŻY'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Saldo końcowe H1 w roku sprzedaży [CHF]'
    ws['B20'] = "=INDEX('05_Harmonogram_roczny'!L:L,$B$10)"
    set_cell_style(ws['A20'])
    set_cell_style(ws['B20'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A21'] = 'Saldo końcowe H2 w roku sprzedaży [CHF]'
    ws['B21'] = "=INDEX('05_Harmonogram_roczny'!M:M,$B$10)"
    set_cell_style(ws['A21'])
    set_cell_style(ws['B21'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A22'] = 'Saldo kredytu łącznie [CHF]'
    ws['B22'] = '=B20+B21'
    set_cell_style(ws['A22'])
    set_cell_style(ws['B22'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Rozliczenie sprzedaży
    ws['A24'] = 'ROZLICZENIE SPRZEDAŻY'
    set_cell_style(ws['A24'], font_bold=True, font_size=12, border=False)
    
    ws['A26'] = 'Cena sprzedaży brutto [CHF]'
    ws['B26'] = '=B15'
    set_cell_style(ws['A26'])
    set_cell_style(ws['B26'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A27'] = 'Koszty sprzedaży [CHF]'
    ws['B27'] = '=B16'
    set_cell_style(ws['A27'])
    set_cell_style(ws['B27'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A28'] = 'Przychód netto po kosztach sprzedaży [CHF]'
    ws['B28'] = '=B26-B27'
    set_cell_style(ws['A28'])
    set_cell_style(ws['B28'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A29'] = 'Spłata kredytu (H1 + H2) [CHF]'
    ws['B29'] = '=B22'
    set_cell_style(ws['A29'])
    set_cell_style(ws['B29'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A30'] = 'Środki po spłacie kredytu [CHF]'
    ws['B30'] = '=B28-B29'
    set_cell_style(ws['A30'], font_bold=True)
    set_cell_style(ws['B30'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    # Analiza zwrotu
    ws['A32'] = 'ANALIZA ZWROTU Z INWESTYCJI'
    set_cell_style(ws['A32'], font_bold=True, font_size=12, border=False)
    
    ws['A34'] = 'Zysk brutto względem wkładu własnego [CHF]'
    ws['B34'] = '=B30-B5'
    set_cell_style(ws['A34'], font_bold=True)
    set_cell_style(ws['B34'], font_bold=True, bg_color='FFEB9C', number_format='#,##0.00')
    
    ws['A35'] = 'ROI względem wkładu własnego'
    ws['B35'] = '=B34/B5'
    set_cell_style(ws['A35'], font_bold=True)
    set_cell_style(ws['B35'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A36'] = 'Średni roczny zwrot (CAGR)'
    ws['B36'] = '=(1+B35)^(1/$B$7)-1'
    set_cell_style(ws['A36'], font_bold=True)
    set_cell_style(ws['B36'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A38'] = 'Komentarz'
    ws['B38'] = '=IF(B35>0,"Zysk na inwestycji","Strata na inwestycji")'
    set_cell_style(ws['A38'], font_bold=True)
    set_cell_style(ws['B38'], font_bold=True, bg_color='FFEB9C')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30


# ============================================================================
# 2) NOWA FUNKCJA - wkleić POD create_sale_analysis_sheet
# ============================================================================

def create_prd_analysis_sheet(wb):
    """Tworzy arkusz 13_Analiza_PRD - Price-to-Rent Ratio, yield i interpretacja wyceny."""
    ws = wb.create_sheet('13_Analiza_PRD')
    
    # Sekcja parametrów
    ws['A2'] = 'Parametr'
    ws['B2'] = 'Wartość'
    set_cell_style(ws['A2'], font_bold=True, bg_color='E0E0E0')
    set_cell_style(ws['B2'], font_bold=True, bg_color='E0E0E0')
    
    ws['A4'] = 'Cena zakupu [CHF]'
    ws['B4'] = "='01_Wejście'!B4"
    set_cell_style(ws['A4'])
    set_cell_style(ws['B4'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A5'] = 'Miesięczny czynsz referencyjny [CHF]'
    ws['B5'] = "='01_Wejście'!B27"
    set_cell_style(ws['A5'])
    set_cell_style(ws['B5'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A6'] = 'Roczny czynsz referencyjny [CHF]'
    ws['B6'] = '=B5*12'
    set_cell_style(ws['A6'])
    set_cell_style(ws['B6'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A7'] = 'Cena zakupu + koszty transakcyjne [CHF]'
    ws['B7'] = "=B4*(1+'00_Stałe'!B10)"
    set_cell_style(ws['A7'])
    set_cell_style(ws['B7'], bg_color='F2F2F2', number_format='#,##0.00')
    
    # Wskaźniki PRD i yield
    ws['A10'] = 'Wskaźniki wyceny (Price-to-Rent Ratio)'
    set_cell_style(ws['A10'], font_bold=True, font_size=12, border=False)
    
    ws['A12'] = 'Price-to-Rent Ratio (PRD) – bez kosztów'
    ws['B12'] = '=B4/B6'
    set_cell_style(ws['A12'], font_bold=True)
    set_cell_style(ws['B12'], font_bold=True, bg_color='FFEB9C', number_format='0.0')
    
    ws['A13'] = 'Price-to-Rent (z kosztami transakcyjnymi)'
    ws['B13'] = '=B7/B6'
    set_cell_style(ws['A13'])
    set_cell_style(ws['B13'], bg_color='F2F2F2', number_format='0.0')
    
    ws['A14'] = 'Brutto yield (roczny czynsz / cena)'
    ws['B14'] = '=B6/B4'
    set_cell_style(ws['A14'], font_bold=True)
    set_cell_style(ws['B14'], font_bold=True, bg_color='FFEB9C', number_format=FORMAT_PERCENTAGE_00)
    
    ws['A15'] = 'Brutto yield (z kosztami)'
    ws['B15'] = '=B6/B7'
    set_cell_style(ws['A15'])
    set_cell_style(ws['B15'], bg_color='F2F2F2', number_format=FORMAT_PERCENTAGE_00)
    
    # Interpretacja PRD
    ws['A18'] = 'Interpretacja wskaźnika Price-to-Rent'
    set_cell_style(ws['A18'], font_bold=True, font_size=12, border=False)
    
    ws['A20'] = 'Klasyfikacja (na podstawie PRD)'
    ws['B20'] = '=IF(B12<15,"Tanio",IF(B12<20,"Normalnie",IF(B12<25,"Drogo","Bardzo drogo")))'
    set_cell_style(ws['A20'], font_bold=True)
    set_cell_style(ws['B20'], font_bold=True, bg_color='FFEB9C')
    
    # Dodatkowe formatowanie warunkowe dla B20
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Tanio', dxf=DifferentialStyle(fill=green_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Normalnie', dxf=DifferentialStyle(fill=yellow_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Drogo', dxf=DifferentialStyle(fill=orange_fill)))
    ws.conditional_formatting.add('B20',
        Rule(type='containsText', operator='containsText', text='Bardzo drogo', dxf=DifferentialStyle(fill=red_fill)))
    
    # Porównanie z kosztem posiadania
    ws['A23'] = 'Koszt posiadania miesięcznie [CHF]'
    ws['B23'] = "='04_Cashflow'!B8"
    set_cell_style(ws['A23'])
    set_cell_style(ws['B23'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws['A24'] = 'Różnica: czynsz – koszt posiadania [CHF/mies.]'
    ws['B24'] = '=B5-B23'
    set_cell_style(ws['A24'])
    set_cell_style(ws['B24'], bg_color='F2F2F2', number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30


# ============================================================================
# 3) ZAMIENIĆ dotychczasową funkcję main() na poniższą wersję
# ============================================================================

def main():
    """Główna funkcja tworząca cały skoroszyt z 14 arkuszami."""
    print("Tworzenie rozszerzonego kalkulatora nieruchomości w Szwajcarii...")
    
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    print("  -> Tworzenie arkusza 00_Stałe...")
    create_constants_sheet(wb)
    
    print("  -> Tworzenie arkusza 01_Wejście...")
    create_input_sheet(wb)
    
    print("  -> Tworzenie arkusza 02_Finansowanie...")
    create_financing_sheet(wb)
    
    print("  -> Tworzenie arkusza 03_Tragbarkeit...")
    create_tragbarkeit_sheet(wb)
    
    print("  -> Tworzenie arkusza 04_Cashflow...")
    create_cashflow_sheet(wb)
    
    print("  -> Tworzenie arkusza 05_Harmonogram_roczny...")
    create_yearly_schedule_sheet(wb)
    
    print("  -> Tworzenie arkusza 06_Harmonogram_miesieczny...")
    create_monthly_schedule_sheet(wb)
    
    print("  -> Tworzenie arkusza 07_Analiza_ROI...")
    create_roi_sheet(wb)
    
    print("  -> Tworzenie arkusza 08_Symulacja_wzrostu_wartości...")
    create_appreciation_sheet(wb)
    
    print("  -> Tworzenie arkusza 09_Koszt_alternatywny_kapitalu...")
    create_opportunity_cost_sheet(wb)
    
    print("  -> Tworzenie arkusza 10_Rent_vs_Buy_30lat...")
    create_rent_vs_buy_sheet(wb)
    
    print("  -> Tworzenie arkusza 11_Stress_test...")
    create_stress_test_sheet(wb)
    
    print("  -> Tworzenie arkusza 12_Analiza_sprzedaży_po_X_latach...")
    create_sale_analysis_sheet(wb)
    
    print("  -> Tworzenie arkusza 13_Analiza_PRD...")
    create_prd_analysis_sheet(wb)
    
    wb.active = wb['01_Wejście']
    
    filename = 'kalkulator_nieruchomosc_CH.xlsx'
    wb.save(filename)
    
    print(f"\n✅ Plik '{filename}' został utworzony pomyślnie!")
    print("\nStruktura arkuszy:")
    print("  00_Stałe - Parametry ogólne")
    print("  01_Wejście - Dane wejściowe użytkownika")
    print("  02_Finansowanie - Analiza kredytu")
    print("  03_Tragbarkeit - Test zdolności kredytowej")
    print("  04_Cashflow - Rzeczywiste koszty miesięczne")
    print("  05_Harmonogram_roczny - Harmonogram spłat rocznych (30 lat)")
    print("  06_Harmonogram_miesieczny - Harmonogram spłat miesięcznych (360 miesięcy)")
    print("  07_Analiza_ROI - Budowa equity i ROI w czasie")
    print("  08_Symulacja_wzrostu_wartości - 3 scenariusze wzrostu wartości nieruchomości")
    print("  09_Koszt_alternatywny_kapitalu - Porównanie equity z alternatywną inwestycją (ETF)")
    print("  10_Rent_vs_Buy_30lat - Porównanie kupna vs wynajmu w horyzoncie 30 lat")
    print("  11_Stress_test - Szok stóp procentowych a koszty i Tragbarkeit")
    print("  12_Analiza_sprzedaży_po_X_latach - Analiza wyniku sprzedaży po X latach, z uwzględnieniem spłaty kredytu")
    print("  13_Analiza_PRD - Price-to-Rent Ratio, yield i interpretacja wyceny")
    print("\nInstrukcja użytkowania:")
    print("1. Otwórz plik w LibreOffice Calc")
    print("2. Przejdź do arkusza '01_Wejście'")
    print("3. Wypełnij niebieskie pola danymi (w tym dobrowolną amortyzację H1)")
    print("4. W arkuszu 07_Analiza_ROI wpisz szacowany roczny wzrost wartości")
    print("5. W arkuszu 08_Symulacja wpisz 3 scenariusze wzrostu")
    print("6. W arkuszu 09_Koszt_alternatywny wpisz stopę zwrotu ETF")
    print("7. W arkuszu 10_Rent_vs_Buy wpisz wzrost czynszu i kosztów")
    print("8. Arkusz 11_Stress_test pokazuje wpływ wzrostu stóp procentowych")
    print("9. W arkuszu 12_Analiza_sprzedaży wpisz horyzont sprzedaży i wzrost wartości")
    print("10. Arkusz 13_Analiza_PRD pokazuje wskaźniki wyceny i yield")
    print("11. Wyniki pojawią się automatycznie we wszystkich arkuszach")
    print("\nPowodzenia w analizie opłacalności zakupu nieruchomości! 🏠🇨🇭")


if __name__ == '__main__':
    main()
